# generate_ceo_presentations.py
#
# Reads the 62-company, 10-sector dataset from
# BenchmarkIQ_Excellence_Model_Filled.xlsx ("All Sectors" sheet) and
# generates ONE CEO_Excellence_Presentation_<Sector>.xlsx per sector,
# matching the navy/gold formatting of the original Automobile-only file.
#
# Run from the scripts folder:  python generate_ceo_presentations.py
# Input expected at the project root:  BenchmarkIQ_Excellence_Model_Filled.xlsx
# Output written to:  CEO_Presentations\CEO_Excellence_Presentation_<Sector>.xlsx

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_PATH  = r"C:\Users\Jignesh\Desktop\bse_portal\BenchmarkIQ_Excellence_Model_Filled.xlsx"
OUTPUT_DIR  = r"C:\Users\Jignesh\Desktop\bse_portal\CEO_Presentations"

# ── Colors (sampled directly from the existing template) ────────────────
NAVY        = "1F3864"
GOLD        = "C9A84C"
LIGHT_BLUE  = "D9E1F2"
WHITE       = "FFFFFF"
DARK_GREY   = "404040"
DARK_GREEN  = "375623"
LIGHT_GREY  = "F2F2F2"
HEAT_GREEN  = "C6EFCE"
HEAT_YELLOW = "FFEB9C"
HEAT_RED    = "FFC7CE"

RATIO_WEIGHTS = {
    "Net Profit Margin":       0.10,
    "EBITDA Margin":           0.10,
    "ROE":                     0.08,
    "ROCE":                    0.05,
    "Operating Profit Margin": 0.02,
    "Revenue Growth YoY":      0.10,
    "3Y Revenue CAGR":         0.08,
    "NP Growth YoY":           0.05,
    "EPS Growth YoY":          0.02,
    "Asset Turnover":          0.07,
    "Debtor Days":             0.05,
    "Inventory Turnover":      0.08,
    "Debt to Equity":          0.08,
    "Interest Coverage":       0.07,
    "Current Ratio":           0.05,
}

HIGHER_BETTER = {
    "Net Profit Margin": True, "EBITDA Margin": True,
    "ROE": True, "ROCE": True, "Operating Profit Margin": True,
    "Revenue Growth YoY": True, "3Y Revenue CAGR": True,
    "NP Growth YoY": True, "EPS Growth YoY": True,
    "Asset Turnover": True, "Debtor Days": False,
    "Inventory Turnover": True, "Debt to Equity": False,
    "Interest Coverage": True, "Current Ratio": True,
}

CATEGORIES = {
    "Profitability": ["Net Profit Margin", "EBITDA Margin", "ROE", "ROCE", "Operating Profit Margin"],
    "Growth":        ["Revenue Growth YoY", "3Y Revenue CAGR", "NP Growth YoY", "EPS Growth YoY"],
    "Efficiency":    ["Asset Turnover", "Debtor Days", "Inventory Turnover"],
    "Safety":        ["Debt to Equity", "Interest Coverage", "Current Ratio"],
}

RATIO_NAMES = list(RATIO_WEIGHTS.keys())


def tier_label(score):
    if score >= 85:   return "Excellence Leader"
    elif score >= 70: return "High Performer"
    elif score >= 55: return "Above Average"
    elif score >= 40: return "Average"
    elif score >= 25: return "Below Average"
    else:             return "Needs Improvement"


def tier_fill_color(tier):
    return {
        "Excellence Leader":  DARK_GREEN,
        "High Performer":     DARK_GREEN,
        "Above Average":      "538135",
        "Average":            "BF8F00",
        "Below Average":      "C45911",
        "Needs Improvement":  "C00000",
    }.get(tier, DARK_GREY)


def percentile_rank(value, all_values, higher_better=True):
    valid = [v for v in all_values if v is not None]
    if not valid or value is None:
        return 50.0
    if higher_better:
        rank = sum(1 for v in valid if v <= value) / len(valid) * 100
    else:
        rank = sum(1 for v in valid if v >= value) / len(valid) * 100
    return round(rank, 1)


def load_sector_data():
    """Read the All Sectors sheet and group rows by sector."""
    wb = openpyxl.load_workbook(INPUT_PATH, data_only=True)
    ws = wb["All Sectors"]

    headers = [c.value for c in ws[1]]
    sectors = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        record = dict(zip(headers, row))
        sector = record["Sector"]
        sectors.setdefault(sector, []).append(record)

    return sectors


def compute_category_scores(companies):
    """
    companies: list of dicts (one per company in a sector), each containing
    raw ratio values under RATIO_NAMES keys.
    Returns: percentile_scores[company] = {ratio: pct}
             category_scores[company]   = {category: weighted_pct_avg}
    """
    percentile_scores = {}
    for ratio in RATIO_NAMES:
        all_vals = [c.get(ratio) for c in companies]
        for c in companies:
            name = c["Company"]
            val = c.get(ratio)
            pct = percentile_rank(val, all_vals, HIGHER_BETTER[ratio])
            percentile_scores.setdefault(name, {})[ratio] = pct

    category_scores = {}
    for c in companies:
        name = c["Company"]
        category_scores[name] = {}
        for cat, ratios in CATEGORIES.items():
            pct_values = [percentile_scores[name][r] for r in ratios]
            category_scores[name][cat] = round(sum(pct_values) / len(pct_values), 1)

    return percentile_scores, category_scores


def heat_color(pct):
    if pct is None:
        return LIGHT_GREY
    if pct >= 75:
        return HEAT_GREEN
    elif pct >= 50:
        return HEAT_YELLOW
    else:
        return HEAT_RED


def style_title(ws, cell_ref, text, merge_to=None):
    ws[cell_ref] = text
    ws[cell_ref].font = Font(bold=True, size=14, color=WHITE)
    ws[cell_ref].fill = PatternFill("solid", fgColor=NAVY)
    if merge_to:
        ws.merge_cells(f"{cell_ref}:{merge_to}")


def build_cover(wb, sector_name, ranked):
    ws = wb.create_sheet("Cover")
    for col in "ABCDEFGHIJKLMN":
        ws.column_dimensions[col].width = 10

    ws["C9"]  = "BenchmarkIQ"
    ws["C9"].font = Font(bold=True, size=36, color=WHITE)
    ws["C11"] = f"{sector_name.upper()} SECTOR"
    ws["C11"].font = Font(bold=True, size=22, color=GOLD)
    ws["C13"] = "Excellence Model — CEO Presentation"
    ws["C13"].font = Font(size=16, color="BDD7EE")

    for r in range(1, 40):
        for c in range(1, 15):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=NAVY)
    # re-apply text on top of the fill (fills above overwrote font color cells only, values persist)
    ws["C9"].font = Font(bold=True, size=36, color=WHITE)
    ws["C11"].font = Font(bold=True, size=22, color=GOLD)
    ws["C13"].font = Font(size=16, color="BDD7EE")

    top = ranked[0]
    ws["C16"] = f"#1 — {top[0]}   |   Excellence Score: {top[1]:.1f} / 100"
    ws["C16"].font = Font(bold=True, size=13, color=WHITE)
    ws["C18"] = f"{len(ranked)} companies benchmarked across 15 financial ratios"
    ws["C18"].font = Font(size=11, color="BDD7EE")


def build_rankings_dashboard(wb, sector_name, ranked, category_scores):
    ws = wb.create_sheet("Rankings Dashboard")
    headers = ["Rank", "Company", "Excellence Score", "Tier",
               "Profitability", "Growth", "Efficiency", "Safety"]

    ws["A2"] = f"BenchmarkIQ Excellence Rankings — {sector_name} Sector"
    ws["A2"].font = Font(bold=True, size=14, color=WHITE)
    ws["A2"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells("A2:H2")

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = Font(bold=True, color=NAVY)
        cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)

    for i, (company, score) in enumerate(ranked):
        r = 5 + i
        tier = tier_label(score)
        short_name = company.replace(" Limited", "").replace(" India", "").strip()
        cat = category_scores[company]

        ws.cell(row=r, column=1, value=i + 1).font = Font(bold=True, color=GOLD if i < 3 else DARK_GREY)
        ws.cell(row=r, column=2, value=short_name).font = Font(color=DARK_GREY)
        ws.cell(row=r, column=3, value=score).font = Font(bold=True, color=DARK_GREEN)
        tier_cell = ws.cell(row=r, column=4, value=tier)
        tier_cell.font = Font(bold=True, color=WHITE)
        tier_cell.fill = PatternFill("solid", fgColor=tier_fill_color(tier))
        ws.cell(row=r, column=5, value=cat["Profitability"]).font = Font(color=DARK_GREY)
        ws.cell(row=r, column=6, value=cat["Growth"]).font = Font(color=DARK_GREY)
        ws.cell(row=r, column=7, value=cat["Efficiency"]).font = Font(color=DARK_GREY)
        ws.cell(row=r, column=8, value=cat["Safety"]).font = Font(color=DARK_GREY)

    widths = [8, 26, 16, 18, 14, 12, 12, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_heatmap(wb, sector_name, ranked, all_ratios, percentile_scores):
    ws = wb.create_sheet("Ratio Heatmap")
    ws["A2"] = "Ratio Heatmap — All Companies vs All Metrics"
    ws["A2"].font = Font(bold=True, size=13, color=NAVY)

    short_labels = [r.replace(" ", "\n") for r in RATIO_NAMES]
    headers = ["Rank", "Company", "Score"] + short_labels

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = Font(bold=True, size=9, color=NAVY)
        cell.alignment = Alignment(wrap_text=True, horizontal="center")
        cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)

    for c, ratio in enumerate(RATIO_NAMES, start=4):
        pct_cell = ws.cell(row=5, column=c, value=f"{int(RATIO_WEIGHTS[ratio]*100)}%")
        pct_cell.font = Font(size=8, italic=True, color=DARK_GREY)
        pct_cell.fill = PatternFill("solid", fgColor=LIGHT_GREY)

    for i, (company, score) in enumerate(ranked):
        r = 6 + i
        short_name = company.replace(" Limited", "").replace(" India", "").strip()
        ws.cell(row=r, column=1, value=i + 1).font = Font(bold=True)
        ws.cell(row=r, column=2, value=short_name)
        ws.cell(row=r, column=3, value=score).font = Font(bold=True)

        for c, ratio in enumerate(RATIO_NAMES, start=4):
            val = all_ratios[company].get(ratio)
            pct = percentile_scores[company].get(ratio)
            cell = ws.cell(row=r, column=c, value=round(val, 1) if val is not None else None)
            cell.fill = PatternFill("solid", fgColor=heat_color(pct))
            cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 8
    for i in range(4, 4 + len(RATIO_NAMES)):
        ws.column_dimensions[get_column_letter(i)].width = 9


def build_category_deep_dive(wb, sector_name, ranked, all_ratios):
    ws = wb.create_sheet("Category Deep Dive")
    ws["A2"] = "Category Deep Dive — Profitability · Growth · Efficiency · Safety"
    ws["A2"].font = Font(bold=True, size=13, color=NAVY)

    cat_cols = {
        "Profitability": (1, ["Net Profit Margin", "EBITDA Margin", "ROE", "ROCE", "Operating Profit Margin"]),
        "Growth":        (7, ["Revenue Growth YoY", "3Y Revenue CAGR", "NP Growth YoY", "EPS Growth YoY"]),
        "Efficiency":    (13, ["Asset Turnover", "Debtor Days", "Inventory Turnover"]),
        "Safety":        (18, ["Debt to Equity", "Interest Coverage", "Current Ratio"]),
    }

    for cat, (start_col, ratios) in cat_cols.items():
        ws.cell(row=4, column=start_col, value=cat).font = Font(bold=True, color=NAVY)
        ws.merge_cells(start_row=4, start_column=start_col,
                        end_row=4, end_column=start_col + len(ratios))

        ws.cell(row=5, column=start_col, value="Company").font = Font(bold=True, size=9)
        for i, ratio in enumerate(ratios):
            cell = ws.cell(row=5, column=start_col + 1 + i, value=ratio.replace(" ", "\n"))
            cell.font = Font(bold=True, size=8)
            cell.alignment = Alignment(wrap_text=True, horizontal="center")

        for i, (company, _) in enumerate(ranked):
            r = 6 + i
            short_name = company.replace(" Limited", "").replace(" India", "").strip()
            ws.cell(row=r, column=start_col, value=short_name).font = Font(size=9)
            for j, ratio in enumerate(ratios):
                val = all_ratios[company].get(ratio)
                ws.cell(row=r, column=start_col + 1 + j,
                        value=round(val, 1) if val is not None else None).font = Font(size=9)

    ws.column_dimensions["A"].width = 18
    for col in "BCDEF":
        ws.column_dimensions[col].width = 9
    ws.column_dimensions["H"].width = 18
    for col in "IJKL":
        ws.column_dimensions[col].width = 9
    ws.column_dimensions["N"].width = 18
    for col in "OPQ":
        ws.column_dimensions[col].width = 9
    ws.column_dimensions["S"].width = 18
    for col in "TUV":
        ws.column_dimensions[col].width = 9


def build_top5_profiles(wb, sector_name, ranked, all_ratios, percentile_scores):
    ws = wb.create_sheet("Top 5 Profiles")
    ws["A2"] = "Top 5 Company Profiles — Detailed Ratio Breakdown"
    ws["A2"].font = Font(bold=True, size=13, color=NAVY)

    top5 = ranked[:5]
    block_starts = [1, 5, 9, 13, 17]

    for idx, (company, score) in enumerate(top5):
        start_col = block_starts[idx]
        short_name = company.replace(" Limited", "").replace(" India", "").strip()
        end_col = start_col + 2

        ws.cell(row=4, column=start_col, value=f"#{idx+1}  {short_name}").font = Font(bold=True, color=NAVY)
        ws.merge_cells(start_row=4, start_column=start_col, end_row=4, end_column=end_col)

        ws.cell(row=5, column=start_col, value=f"Excellence Score: {score:.1f}").font = Font(italic=True, size=9)
        ws.merge_cells(start_row=5, start_column=start_col, end_row=5, end_column=end_col)

        for c, label in zip([start_col, start_col + 1, start_col + 2], ["Metric", "Value", "Pct"]):
            cell = ws.cell(row=6, column=c, value=label)
            cell.font = Font(bold=True, size=9)
            cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)

        for i, ratio in enumerate(RATIO_NAMES):
            r = 7 + i
            val = all_ratios[company].get(ratio)
            pct = percentile_scores[company].get(ratio)
            ws.cell(row=r, column=start_col, value=ratio).font = Font(size=8)
            ws.cell(row=r, column=start_col + 1,
                    value=round(val, 1) if val is not None else None).font = Font(size=8)
            ws.cell(row=r, column=start_col + 2,
                    value=f"{pct:.1f}th" if pct is not None else "N/A").font = Font(size=8)

    for col_idx in range(1, 20):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12


def build_methodology(wb, sector_name, n_companies):
    ws = wb.create_sheet("Methodology")
    ws["A2"] = "Methodology — How Excellence Scores Are Calculated"
    ws["A2"].font = Font(bold=True, size=13, color=NAVY)

    lines = [
        (4, "OVERVIEW", True),
        (5, f"The Excellence Model ranks {n_companies} Indian {sector_name.lower()} companies using "
            f"{len(RATIO_NAMES)} financial ratios across 4 categories.", False),
        (6, "Each company receives a percentile rank (0-100) per ratio, then a weighted average score is computed.", False),
        (8, "SCORING METHODOLOGY", True),
        (9, "Step 1: Calculate 15 ratios from P&L and Balance Sheet data for each company (latest FY).", False),
        (10, f"Step 2: Rank each company's ratio against all {n_companies} peers using percentile ranking.", False),
        (11, "Step 3: Apply category weights and compute a weighted Excellence Score (0-100).", False),
        (13, "TIER CLASSIFICATION", True),
        (14, "85-100  →  Excellence Leader   |  70-84  →  High Performer   |  55-69  →  Above Average", False),
        (15, "40-54   →  Average              |  25-39  →  Below Average    |  0-24   →  Needs Improvement", False),
        (17, "DATA SOURCE", True),
        (18, "BSE/NSE filings via Screener.in  |  Stored in MySQL  |  FY ending March 2026", False),
        (20, "COLOR CODING (HEATMAP)", True),
        (21, "Green  =  Top quartile (75th percentile+)", False),
        (22, "Yellow =  Middle (50th-75th percentile)", False),
        (23, "Red    =  Bottom half (below 50th percentile)", False),
        (25, "RATIO WEIGHTS", True),
    ]
    for r, text, bold in lines:
        cell = ws.cell(row=r, column=1, value=text)
        cell.font = Font(bold=bold, size=12 if bold and r == 2 else 10, color=NAVY if bold else DARK_GREY)

    cat_map = {r: cat for cat, ratios in CATEGORIES.items() for r in ratios}
    for i, ratio in enumerate(RATIO_NAMES):
        r = 26 + i
        cat = cat_map[ratio]
        pct = int(RATIO_WEIGHTS[ratio] * 100)
        ws.cell(row=r, column=1, value=f"  {ratio}  ({cat})  →  {pct}%").font = Font(size=9, color=DARK_GREY)

    ws.column_dimensions["A"].width = 100


def generate_sector_workbook(sector_name, companies):
    """companies: list of raw record dicts for this sector (from All Sectors sheet)."""
    percentile_scores, category_scores = compute_category_scores(companies)

    all_ratios = {c["Company"]: {r: c.get(r) for r in RATIO_NAMES} for c in companies}
    scored = [(c["Company"], c["Excellence Score"]) for c in companies]
    ranked = sorted(scored, key=lambda x: x[1], reverse=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_cover(wb, sector_name, ranked)
    build_rankings_dashboard(wb, sector_name, ranked, category_scores)
    build_heatmap(wb, sector_name, ranked, all_ratios, percentile_scores)
    build_category_deep_dive(wb, sector_name, ranked, all_ratios)
    build_top5_profiles(wb, sector_name, ranked, all_ratios, percentile_scores)
    build_methodology(wb, sector_name, len(companies))

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    safe_name = sector_name.replace(" & ", "_").replace(" ", "_")
    out_path = os.path.join(OUTPUT_DIR, f"CEO_Excellence_Presentation_{safe_name}.xlsx")
    wb.save(out_path)
    return out_path, len(companies)


def main():
    if not os.path.exists(INPUT_PATH):
        print(f"ERROR: Input file not found at {INPUT_PATH}")
        print("Make sure BenchmarkIQ_Excellence_Model_Filled.xlsx is at the project root.")
        return

    sectors = load_sector_data()
    print(f"Found {len(sectors)} sectors, {sum(len(v) for v in sectors.values())} companies total")
    print("=" * 60)

    for sector_name, companies in sorted(sectors.items()):
        if len(companies) < 2:
            print(f"SKIPPED {sector_name} — only {len(companies)} company, need at least 2 for percentiles")
            continue
        out_path, n = generate_sector_workbook(sector_name, companies)
        print(f"OK: {sector_name:<25} ({n} companies) -> {os.path.basename(out_path)}")

    print("=" * 60)
    print(f"All workbooks saved to: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()