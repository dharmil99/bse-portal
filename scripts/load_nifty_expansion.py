import os
import sys
sys.path.insert(0, r'C:\Users\Jignesh\Desktop\bse_portal')

from openpyxl import load_workbook
from sqlalchemy import text
from scripts.db_connect import get_engine

engine = get_engine()
RAW_DIR = r'C:\Users\Jignesh\Desktop\bse_portal\data\raw'

# ── New companies to add ──────────────────────────────
NEW_COMPANIES = [
    # Banking / Financial Services
    {"file": "AxisBank",       "bse": "532215", "nse": "AXISBANK",    "name": "Axis Bank Limited",            "sector": "Banking",        "industry": "Private Bank",       "mcap": 350000},
    {"file": "BajajFinance",   "bse": "500034", "nse": "BAJFINANCE",  "name": "Bajaj Finance Limited",        "sector": "Banking",        "industry": "NBFC",               "mcap": 420000},
    {"file": "BajajFinserv",   "bse": "532978", "nse": "BAJAJFINSV",  "name": "Bajaj Finserv Limited",        "sector": "Banking",        "industry": "Financial Services", "mcap": 280000},
    {"file": "HDFCLife",       "bse": "540777", "nse": "HDFCLIFE",    "name": "HDFC Life Insurance",          "sector": "Banking",        "industry": "Insurance",          "mcap": 140000},
    {"file": "SBILife",        "bse": "540719", "nse": "SBILIFE",     "name": "SBI Life Insurance",           "sector": "Banking",        "industry": "Insurance",          "mcap": 160000},
    {"file": "ShriramFinance", "bse": "511218", "nse": "SHRIRAMFIN",  "name": "Shriram Finance Limited",      "sector": "Banking",        "industry": "NBFC",               "mcap": 95000},
    # FMCG
    {"file": "TataConsumer",   "bse": "500800", "nse": "TATACONSUM",  "name": "Tata Consumer Products",       "sector": "FMCG",           "industry": "Consumer Goods",     "mcap": 95000},
    # Healthcare / Pharma
    {"file": "ApolloHospitals","bse": "508869", "nse": "APOLLOHOSP",  "name": "Apollo Hospitals Enterprise",  "sector": "Pharmaceuticals","industry": "Hospitals",          "mcap": 95000},
    {"file": "MaxHealthcare",  "bse": "543271", "nse": "MAXHEALTH",   "name": "Max Healthcare Institute",     "sector": "Pharmaceuticals","industry": "Hospitals",          "mcap": 85000},
]

def excel_date_to_year(val):
    try:
        from datetime import datetime, timedelta
        if isinstance(val, (int, float)):
            dt = datetime(1899, 12, 30) + timedelta(days=int(val))
        else:
            dt = val
        fy = dt.year + 1 if dt.month >= 4 else dt.year
        return f"FY{str(fy)[2:]}"
    except:
        return None

def load_data_sheet(wb):
    ws = wb['Data Sheet']
    data = {}
    current_section = None
    for row in ws.iter_rows(values_only=True):
        if not any(row):
            continue
        label = str(row[0]).strip() if row[0] else ""
        if label == "PROFIT & LOSS":      current_section = "PL";   continue
        elif label == "BALANCE SHEET":    current_section = "BS";   continue
        elif label == "CASH FLOW:":       current_section = "CF";   continue
        elif label == "Quarters":         current_section = "Q";    continue
        elif label == "META":             current_section = "META"; continue
        values = [v for v in row[1:] if v is not None]
        if values and current_section:
            data[f"{current_section}_{label}"] = values
    return data

def get_or_create_sector(conn, sector_name, industry):
    conn.execute(text(
        "INSERT IGNORE INTO sectors (sector_name, industry) VALUES (:s, :i)"
    ), {"s": sector_name, "i": industry})
    conn.commit()
    row = conn.execute(text(
        "SELECT sector_id FROM sectors WHERE sector_name = :s"
    ), {"s": sector_name}).fetchone()
    return row[0]

def load_company(co):
    filepath = os.path.join(RAW_DIR, co["file"] + ".xlsx")
    if not os.path.exists(filepath):
        print(f"  ❌ File not found: {co['file']}.xlsx")
        return False

    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"  ❌ Cannot open: {e}")
        return False

    if 'Data Sheet' not in wb.sheetnames:
        print(f"  ❌ No Data Sheet tab in {co['file']}.xlsx")
        return False

    data = load_data_sheet(wb)

    with engine.connect() as conn:
        # Get or create sector
        sector_id = get_or_create_sector(conn, co["sector"], co["industry"])

        # Insert company
        conn.execute(text("""
            INSERT IGNORE INTO companies
                (bse_code, nse_symbol, company_name, sector_id, market_cap)
            VALUES (:bse, :nse, :name, :sid, :mcap)
        """), {
            "bse": co["bse"], "nse": co["nse"],
            "name": co["name"], "sid": sector_id,
            "mcap": co["mcap"]
        })
        conn.commit()

        result = conn.execute(text(
            "SELECT company_id FROM companies WHERE bse_code = :b"
        ), {"b": co["bse"]}).fetchone()
        company_id = result[0]
        print(f"  ✅ Company ID: {company_id} — {co['name']}")

        # Get years
        years_raw = data.get("PL_Report Date", [])
        years = [excel_date_to_year(y) for y in years_raw]
        print(f"  📅 Years: {[y for y in years if y]}")

        # ── P&L ──────────────────────────────────────
        sales        = data.get("PL_Sales", [])
        raw_mat      = data.get("PL_Raw Material Cost", [None]*12)
        emp_cost     = data.get("PL_Employee Cost", [None]*12)
        other_income = data.get("PL_Other Income", [None]*12)
        depreciation = data.get("PL_Depreciation", [None]*12)
        interest     = data.get("PL_Interest", [None]*12)
        pbt          = data.get("PL_Profit before tax", [None]*12)
        tax          = data.get("PL_Tax", [None]*12)
        net_profit   = data.get("PL_Net profit", [None]*12)
        dividend     = data.get("PL_Dividend Amount", [None]*12)

        for i, fy in enumerate(years):
            if not fy or i >= len(sales):
                continue
            try:
                conn.execute(text("""
                    INSERT INTO profit_loss
                        (company_id, fiscal_year, sales, raw_material,
                         employee_cost, other_income, depreciation,
                         interest, profit_before_tax, tax,
                         net_profit, dividend_amount)
                    VALUES
                        (:cid,:fy,:s,:rm,:emp,:oi,:dep,:int,:pbt,:tax,:np,:div)
                    ON DUPLICATE KEY UPDATE
                        sales=VALUES(sales),
                        net_profit=VALUES(net_profit)
                """), {
                    "cid": company_id, "fy": fy,
                    "s":   sales[i]        if i < len(sales)        else None,
                    "rm":  raw_mat[i]      if i < len(raw_mat)      else None,
                    "emp": emp_cost[i]     if i < len(emp_cost)     else None,
                    "oi":  other_income[i] if i < len(other_income) else None,
                    "dep": depreciation[i] if i < len(depreciation) else None,
                    "int": interest[i]     if i < len(interest)     else None,
                    "pbt": pbt[i]          if i < len(pbt)          else None,
                    "tax": tax[i]          if i < len(tax)          else None,
                    "np":  net_profit[i]   if i < len(net_profit)   else None,
                    "div": dividend[i]     if i < len(dividend)     else None,
                })
            except Exception as e:
                print(f"    ⚠️ P&L error {fy}: {e}")

        # ── Balance Sheet ─────────────────────────────
        bs_years = [excel_date_to_year(y) for y in data.get("BS_Report Date", [])]
        eq_cap   = data.get("BS_Equity Share Capital", [])
        reserves = data.get("BS_Reserves", [])
        borrow   = data.get("BS_Borrowings", [])
        oth_lib  = data.get("BS_Other Liabilities", [])
        tot_lib  = data.get("BS_Total", [])
        net_blk  = data.get("BS_Net Block", [])
        cwip     = data.get("BS_Capital Work in Progress", [])
        invest   = data.get("BS_Investments", [])
        oth_ast  = data.get("BS_Other Assets", [])
        recv     = data.get("BS_Receivables", [])
        inv      = data.get("BS_Inventory", [])
        cash     = data.get("BS_Cash & Bank", [])

        for i, fy in enumerate(bs_years):
            if not fy:
                continue
            try:
                conn.execute(text("""
                    INSERT INTO balance_sheet
                        (company_id, fiscal_year, equity_capital,
                         reserves, borrowings, other_liabilities,
                         total_liabilities, net_block, cwip,
                         investments, other_assets, receivables,
                         inventory, cash_and_bank)
                    VALUES
                        (:cid,:fy,:eq,:res,:bor,:olib,:tlib,
                         :nb,:cwip,:inv,:oast,:recv,:invt,:cash)
                    ON DUPLICATE KEY UPDATE
                        reserves=VALUES(reserves),
                        borrowings=VALUES(borrowings)
                """), {
                    "cid":  company_id, "fy": fy,
                    "eq":   eq_cap[i]  if i < len(eq_cap)  else None,
                    "res":  reserves[i]if i < len(reserves) else None,
                    "bor":  borrow[i]  if i < len(borrow)  else None,
                    "olib": oth_lib[i] if i < len(oth_lib) else None,
                    "tlib": tot_lib[i] if i < len(tot_lib) else None,
                    "nb":   net_blk[i] if i < len(net_blk) else None,
                    "cwip": cwip[i]    if i < len(cwip)    else None,
                    "inv":  invest[i]  if i < len(invest)  else None,
                    "oast": oth_ast[i] if i < len(oth_ast) else None,
                    "recv": recv[i]    if i < len(recv)    else None,
                    "invt": inv[i]     if i < len(inv)     else None,
                    "cash": cash[i]    if i < len(cash)    else None,
                })
            except Exception as e:
                print(f"    ⚠️ BS error {fy}: {e}")

        # ── Cash Flow ─────────────────────────────────
        cf_years = [excel_date_to_year(y) for y in data.get("CF_Report Date", [])]
        op_cf    = data.get("CF_Cash from Operating Activity", [])
        inv_cf   = data.get("CF_Cash from Investing Activity", [])
        fin_cf   = data.get("CF_Cash from Financing Activity", [])
        net_cf   = data.get("CF_Net Cash Flow", [])

        for i, fy in enumerate(cf_years):
            if not fy:
                continue
            try:
                conn.execute(text("""
                    INSERT INTO cash_flow
                        (company_id, fiscal_year, operating_cf,
                         investing_cf, financing_cf, net_cash_flow)
                    VALUES (:cid,:fy,:ocf,:icf,:fcf,:ncf)
                    ON DUPLICATE KEY UPDATE
                        operating_cf=VALUES(operating_cf)
                """), {
                    "cid": company_id, "fy": fy,
                    "ocf": op_cf[i]  if i < len(op_cf)  else None,
                    "icf": inv_cf[i] if i < len(inv_cf) else None,
                    "fcf": fin_cf[i] if i < len(fin_cf) else None,
                    "ncf": net_cf[i] if i < len(net_cf) else None,
                })
            except Exception as e:
                print(f"    ⚠️ CF error {fy}: {e}")

        # ── Quarterly Results ─────────────────────────
        q_years_raw = data.get("PL_Report Date", [])
        q_years     = [excel_date_to_year(y) for y in q_years_raw]

        for i, fy in enumerate(q_years):
            if not fy or i >= len(sales):
                continue
            try:
                rev = sales[i]
                np  = net_profit[i] if i < len(net_profit) else None
                dep = depreciation[i] if i < len(depreciation) else None
                int_= interest[i] if i < len(interest) else None
                ebitda = None
                if np and dep and int_:
                    try:
                        ebitda = float(np) + float(dep) + float(int_)
                    except:
                        pass
                quarter = f"Q4{fy}"
                period  = f"20{fy[2:]}-03-31"

                conn.execute(text("""
                    INSERT INTO quarterly_results
                        (company_id, quarter, period_end,
                         revenue, net_profit, ebitda)
                    VALUES (:cid,:q,:p,:r,:np,:eb)
                    ON DUPLICATE KEY UPDATE
                        revenue=VALUES(revenue),
                        net_profit=VALUES(net_profit)
                """), {
                    "cid": company_id, "q": quarter,
                    "p": period, "r": rev,
                    "np": np, "eb": ebitda
                })
            except Exception as e:
                print(f"    ⚠️ QR error {fy}: {e}")

        conn.commit()
        print(f"  ✅ All data loaded for {co['name']}")
        return True

# ── Main ──────────────────────────────────────────────
if __name__ == "__main__":
    print("📊 Loading 9 new Nifty 50 companies")
    print("=" * 50)
    loaded = 0
    for co in NEW_COMPANIES:
        print(f"\nLoading: {co['name']}")
        if load_company(co):
            loaded += 1
    print(f"\n{'='*50}")
    print(f"✅ Done! {loaded}/9 companies loaded")