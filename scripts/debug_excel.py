import sys
import os
import openpyxl
import pandas as pd
from sqlalchemy import text

sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from db_connect import get_engine

# ── Company list ─────────────────────────────────────────────────────────────
COMPANIES = [
    {"file": "Infosys",          "bse": "500209", "nse": "INFY",        "name": "Infosys Limited",               "sector": "Information Technology", "industry": "IT Services",    "mcap": 621000.00},
    {"file": "TCS",              "bse": "532540", "nse": "TCS",         "name": "Tata Consultancy Services",     "sector": "Information Technology", "industry": "IT Services",    "mcap": 1380000.00},
    {"file": "Wipro",            "bse": "507685", "nse": "WIPRO",       "name": "Wipro Limited",                 "sector": "Information Technology", "industry": "IT Services",    "mcap": 260000.00},
    {"file": "HCLTech",          "bse": "532281", "nse": "HCLTECH",     "name": "HCL Technologies Limited",      "sector": "Information Technology", "industry": "IT Services",    "mcap": 390000.00},
    {"file": "TechMahindra",     "bse": "532755", "nse": "TECHM",       "name": "Tech Mahindra Limited",         "sector": "Information Technology", "industry": "IT Services",    "mcap": 120000.00},
    {"file": "SunPharma",        "bse": "524715", "nse": "SUNPHARMA",   "name": "Sun Pharmaceutical Industries", "sector": "Pharmaceuticals",        "industry": "Pharma",         "mcap": 360000.00},
    {"file": "DrReddys",         "bse": "500124", "nse": "DRREDDY",     "name": "Dr. Reddys Laboratories",       "sector": "Pharmaceuticals",        "industry": "Pharma",         "mcap": 95000.00},
    {"file": "Cipla",            "bse": "500087", "nse": "CIPLA",       "name": "Cipla Limited",                 "sector": "Pharmaceuticals",        "industry": "Pharma",         "mcap": 100000.00},
    {"file": "DivisLab",         "bse": "532488", "nse": "DIVISLAB",    "name": "Divis Laboratories",            "sector": "Pharmaceuticals",        "industry": "Pharma",         "mcap": 130000.00},
    {"file": "HDFCBank",         "bse": "500180", "nse": "HDFCBANK",    "name": "HDFC Bank Limited",             "sector": "Banking",                "industry": "Private Bank",   "mcap": 1100000.00},
    {"file": "ICICIBank",        "bse": "532174", "nse": "ICICIBANK",   "name": "ICICI Bank Limited",            "sector": "Banking",                "industry": "Private Bank",   "mcap": 780000.00},
    {"file": "KotakBank",        "bse": "500247", "nse": "KOTAKBANK",   "name": "Kotak Mahindra Bank",           "sector": "Banking",                "industry": "Private Bank",   "mcap": 355000.00},
    {"file": "SBI",              "bse": "500112", "nse": "SBIN",        "name": "State Bank of India",           "sector": "Banking",                "industry": "Public Bank",    "mcap": 700000.00},
    {"file": "HUL",              "bse": "500696", "nse": "HINDUNILVR",  "name": "Hindustan Unilever Limited",    "sector": "FMCG",                   "industry": "Consumer Goods", "mcap": 560000.00},
    {"file": "ITC",              "bse": "500875", "nse": "ITC",         "name": "ITC Limited",                   "sector": "FMCG",                   "industry": "Consumer Goods", "mcap": 560000.00},
    {"file": "NestleIndia",      "bse": "500790", "nse": "NESTLEIND",   "name": "Nestle India Limited",          "sector": "FMCG",                   "industry": "Consumer Goods", "mcap": 230000.00},
    {"file": "Britannia",        "bse": "500825", "nse": "BRITANNIA",   "name": "Britannia Industries",          "sector": "FMCG",                   "industry": "Consumer Goods", "mcap": 120000.00},
    {"file": "TataMotors",       "bse": "500570", "nse": "TATAMOTORS",  "name": "Tata Motors Limited",           "sector": "Automobile",             "industry": "Auto OEM",       "mcap": 340000.00},
    {"file": "Maruti",           "bse": "532500", "nse": "MARUTI",      "name": "Maruti Suzuki India",           "sector": "Automobile",             "industry": "Auto OEM",       "mcap": 380000.00},
    {"file": "MahindraMahindra", "bse": "500520", "nse": "M&M",         "name": "Mahindra and Mahindra",         "sector": "Automobile",             "industry": "Auto OEM",       "mcap": 290000.00},
    {"file": "BajajAuto",        "bse": "532977", "nse": "BAJAJ-AUTO",  "name": "Bajaj Auto Limited",            "sector": "Automobile",             "industry": "Auto OEM",       "mcap": 230000.00},
]

RAW_DIR = r"C:\Users\Jignesh\Desktop\bse_portal\data\raw"

# ── Excel parser: reads from "Data Sheet" tab ─────────────────────────────────
def parse_screener_excel(filepath):
    """
    Screener.in Excel files have a hidden 'Data Sheet' tab with clean data.
    Row 15 = Report Date (datetime objects for each year)
    Row 16 onwards = financial line items with actual values
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Check if Data Sheet exists (new format)
    if "Data Sheet" in wb.sheetnames:
        return parse_data_sheet(wb)
    else:
        # Fallback: old format where data is in first sheet
        return parse_legacy_sheet(wb)


def parse_data_sheet(wb):
    """Parse the modern Screener.in Excel format with Data Sheet tab."""
    ws = wb["Data Sheet"]
    all_rows = list(ws.iter_rows(values_only=True))

    # Find the P&L Report Date row (has datetime values)
    year_row_idx = None
    year_cols = []
    year_labels = []

    for i, row in enumerate(all_rows):
        if row[0] == "Report Date" and i < 40:  # P&L section only
            year_row_idx = i
            for j, val in enumerate(row[1:], start=1):
                if hasattr(val, 'year'):  # it's a datetime
                    year_labels.append(str(val.year))
                    year_cols.append(j)
            break

    if not year_cols:
        raise ValueError("Could not find Report Date row with year values")

    # Build a dict: row_label -> {year: value}
    data = {}
    for row in all_rows[year_row_idx + 1:]:
        label = row[0]
        if label is None:
            continue
        label = str(label).strip()
        values = {}
        for col_idx, year in zip(year_cols, year_labels):
            val = row[col_idx] if col_idx < len(row) else None
            try:
                values[year] = float(val) if val is not None else None
            except (TypeError, ValueError):
                values[year] = None
        data[label] = values

    return data, year_labels


def parse_legacy_sheet(wb):
    """Fallback for old Screener.in format (like Infosys.xlsx)."""
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))

    # Find Narration row
    header_row_idx = None
    for i, row in enumerate(all_rows):
        if str(row[0]).strip() == "Narration":
            header_row_idx = i
            break

    if header_row_idx is None:
        raise ValueError("Could not find Narration row")

    # Get year columns
    year_cols = []
    year_labels = []
    for j, val in enumerate(all_rows[header_row_idx][1:], start=1):
        if val is None:
            continue
        s = str(val)
        if s.startswith("201") or s.startswith("202"):
            year_labels.append(s[:4])
            year_cols.append(j)
        elif hasattr(val, 'year'):
            year_labels.append(str(val.year))
            year_cols.append(j)

    if not year_cols:
        raise ValueError("No year columns found in legacy sheet")

    data = {}
    for row in all_rows[header_row_idx + 1:]:
        label = row[0]
        if label is None:
            continue
        label = str(label).strip()
        values = {}
        for col_idx, year in zip(year_cols, year_labels):
            val = row[col_idx] if col_idx < len(row) else None
            try:
                values[year] = float(val) if val is not None else None
            except (TypeError, ValueError):
                values[year] = None
        data[label] = values

    return data, year_labels


def safe_get(data, label, year):
    """Safely get a value from parsed data dict."""
    try:
        val = data.get(label, {}).get(year)
        return float(val) if val is not None else None
    except (TypeError, ValueError):
        return None


# ── Database helpers ──────────────────────────────────────────────────────────
def get_or_create_sector(conn, sector_name, industry):
    conn.execute(text(
        "INSERT IGNORE INTO sectors (sector_name, industry) VALUES (:s, :i)"
    ), {"s": sector_name, "i": industry})
    conn.commit()
    row = conn.execute(text(
        "SELECT sector_id FROM sectors WHERE sector_name = :s"
    ), {"s": sector_name}).fetchone()
    return row[0]


def get_or_create_company(conn, bse, nse, name, sector_id, mcap):
    conn.execute(text(
        "INSERT IGNORE INTO companies "
        "(bse_code, nse_symbol, company_name, sector_id, market_cap) "
        "VALUES (:b, :n, :nm, :s, :m)"
    ), {"b": bse, "n": nse, "nm": name, "s": sector_id, "m": mcap})
    conn.commit()
    row = conn.execute(text(
        "SELECT company_id FROM companies WHERE bse_code = :b"
    ), {"b": bse}).fetchone()
    return row[0]


# ── Load one company ──────────────────────────────────────────────────────────
def load_company(conn, company_info):
    filepath = os.path.join(RAW_DIR, company_info["file"] + ".xlsx")

    if not os.path.exists(filepath):
        print("  SKIP: File not found -> " + filepath)
        return False

    try:
        data, years = parse_screener_excel(filepath)
    except Exception as e:
        print("  ERROR parsing " + company_info["file"] + ": " + str(e))
        return False

    if not years:
        print("  ERROR: No years found in " + company_info["file"])
        return False

    sector_id  = get_or_create_sector(conn, company_info["sector"], company_info["industry"])
    company_id = get_or_create_company(conn, company_info["bse"], company_info["nse"],
                                       company_info["name"], sector_id, company_info["mcap"])

    rows_inserted = 0
    for year in years:
        try:
            revenue      = safe_get(data, "Sales", year)
            net_profit   = safe_get(data, "Net profit", year)
            ebitda       = safe_get(data, "Operating Profit", year)
            eps_val      = safe_get(data, "EPS", year)
            total_debt   = safe_get(data, "Borrowings", year)
            equity       = safe_get(data, "Equity Share Capital", year)
            total_assets = safe_get(data, "Total", year)

            quarter = "Q4FY" + str(year)[2:]
            period  = str(year) + "-03-31"

            conn.execute(text("""
                INSERT INTO quarterly_results
                (company_id, quarter, period_end, revenue, net_profit,
                 ebitda, total_debt, equity, total_assets, eps)
                VALUES (:cid, :q, :p, :r, :np, :eb, :td, :eq, :ta, :eps)
                ON DUPLICATE KEY UPDATE
                    revenue=VALUES(revenue),
                    net_profit=VALUES(net_profit),
                    ebitda=VALUES(ebitda),
                    total_debt=VALUES(total_debt),
                    equity=VALUES(equity),
                    total_assets=VALUES(total_assets),
                    eps=VALUES(eps)
            """), {
                "cid": company_id, "q": quarter, "p": period,
                "r": revenue,      "np": net_profit, "eb": ebitda,
                "td": total_debt,  "eq": equity,     "ta": total_assets,
                "eps": eps_val
            })
            rows_inserted += 1
        except Exception as e:
            print("  WARNING year " + str(year) + ": " + str(e))

    conn.commit()
    print("  OK: " + company_info["name"] + " -> " + str(rows_inserted) + " years loaded")
    return True


# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("BSE Portal - Bulk Company Loader")
    print("=" * 55)

    try:
        engine = get_engine()
        conn   = engine.connect()
        print("Connected to MySQL\n")
    except Exception as e:
        print("ERROR: Cannot connect to MySQL: " + str(e))
        sys.exit(1)

    loaded = 0
    skipped = 0

    for co in COMPANIES:
        print("Loading: " + co["name"])
        if load_company(conn, co):
            loaded += 1
        else:
            skipped += 1

    conn.close()
    print("\n" + "=" * 55)
    print("DONE!  Loaded: " + str(loaded) + "   Skipped: " + str(skipped))

    if skipped > 0:
        print("\nFor skipped companies — download their Excel from")
        print("screener.in and save to: " + RAW_DIR)