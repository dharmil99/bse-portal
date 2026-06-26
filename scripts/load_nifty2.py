import os
import sys
sys.path.insert(0, r'C:\Users\Jignesh\Desktop\bse_portal')

from openpyxl import load_workbook
from sqlalchemy import text
from scripts.db_connect import get_engine

engine = get_engine()
RAW_DIR = r'C:\Users\Jignesh\Desktop\bse_portal\data\raw'

NEW_COMPANIES = [
    # Energy
    {"file": "Reliance Industries", "bse": "500325", "nse": "RELIANCE",   "name": "Reliance Industries Limited",  "sector": "Energy",             "industry": "Oil & Gas",        "mcap": 1700000},
    {"file": "O N G C",             "bse": "500312", "nse": "ONGC",       "name": "Oil and Natural Gas Corp",     "sector": "Energy",             "industry": "Oil & Gas",        "mcap": 350000},
    {"file": "Coal India",          "bse": "533278", "nse": "COALINDIA",  "name": "Coal India Limited",           "sector": "Energy",             "industry": "Mining",           "mcap": 250000},
    {"file": "NTPC",                "bse": "532555", "nse": "NTPC",       "name": "NTPC Limited",                 "sector": "Energy",             "industry": "Power",            "mcap": 320000},
    {"file": "Power Grid Corpn",    "bse": "532898", "nse": "POWERGRID",  "name": "Power Grid Corporation",       "sector": "Energy",             "industry": "Power",            "mcap": 280000},
    # Metals
    {"file": "Tata Steel",          "bse": "500470", "nse": "TATASTEEL",  "name": "Tata Steel Limited",           "sector": "Metals & Mining",    "industry": "Steel",            "mcap": 180000},
    {"file": "JSW Steel",           "bse": "500228", "nse": "JSWSTEEL",   "name": "JSW Steel Limited",            "sector": "Metals & Mining",    "industry": "Steel",            "mcap": 220000},
    {"file": "Hindalco Inds",       "bse": "500440", "nse": "HINDALCO",   "name": "Hindalco Industries Limited",  "sector": "Metals & Mining",    "industry": "Aluminium",        "mcap": 150000},
    # Cement & Construction
    {"file": "UltraTech Cem",       "bse": "532538", "nse": "ULTRACEMCO", "name": "UltraTech Cement Limited",     "sector": "Cement",             "industry": "Cement",           "mcap": 350000},
    {"file": "Grasim Inds",         "bse": "500300", "nse": "GRASIM",     "name": "Grasim Industries Limited",    "sector": "Cement",             "industry": "Cement",           "mcap": 180000},
    {"file": "Larsen & Toubro",     "bse": "500510", "nse": "LT",         "name": "Larsen and Toubro Limited",    "sector": "Construction",       "industry": "Engineering",      "mcap": 450000},
    # Telecom
    {"file": "Bharti Airtel",       "bse": "532454", "nse": "BHARTIARTL", "name": "Bharti Airtel Limited",        "sector": "Telecom",            "industry": "Telecom",          "mcap": 900000},
    {"file": "Indus Towers",        "bse": "534816", "nse": "INDUSTOWER", "name": "Indus Towers Limited",         "sector": "Telecom",            "industry": "Telecom",          "mcap": 90000},
    # Consumer
    {"file": "Titan Company",       "bse": "500114", "nse": "TITAN",      "name": "Titan Company Limited",        "sector": "Consumer",           "industry": "Jewellery",        "mcap": 320000},
    {"file": "Trent",               "bse": "500251", "nse": "TRENT",      "name": "Trent Limited",                "sector": "Consumer",           "industry": "Retail",           "mcap": 180000},
    {"file": "Eternal",             "bse": "543320", "nse": "ZOMATO",     "name": "Eternal Limited",              "sector": "Consumer",           "industry": "Food Delivery",    "mcap": 220000},
    # Conglomerate
    {"file": "Adani Enterp",        "bse": "512599", "nse": "ADANIENT",   "name": "Adani Enterprises Limited",    "sector": "Conglomerate",       "industry": "Diversified",      "mcap": 320000},
    # Financial
    {"file": "Jio Financial",       "bse": "543771", "nse": "JIOFIN",     "name": "Jio Financial Services",       "sector": "Banking",            "industry": "Financial Services","mcap": 150000},
]

def excel_date_to_year(val):
    try:
        from datetime import datetime, timedelta
        if isinstance(val, (int, float)):
            dt = datetime(1899, 12, 30) + timedelta(days=int(val))
        else:
            dt = val
        fy = dt.year + 1 if dt.month >= 4 else dt.year
        return f"FY{str(fy)[2:]}"
    except:
        return None

def load_data_sheet(wb):
    ws = wb['Data Sheet']
    data = {}
    current_section = None
    for row in ws.iter_rows(values_only=True):
        if not any(row):
            continue
        label = str(row[0]).strip() if row[0] else ""
        if label == "PROFIT & LOSS":      current_section = "PL";   continue
        elif label == "BALANCE SHEET":    current_section = "BS";   continue
        elif label == "CASH FLOW:":       current_section = "CF";   continue
        elif label == "Quarters":         current_section = "Q";    continue
        elif label == "META":             current_section = "META"; continue
        values = [v for v in row[1:] if v is not None]
        if values and current_section:
            data[f"{current_section}_{label}"] = values
    return data

def get_or_create_sector(conn, sector_name, industry):
    conn.execute(text(
        "INSERT IGNORE INTO sectors (sector_name, industry) VALUES (:s, :i)"
    ), {"s": sector_name, "i": industry})
    conn.commit()
    row = conn.execute(text(
        "SELECT sector_id FROM sectors WHERE sector_name = :s AND industry = :i"
    ), {"s": sector_name, "i": industry}).fetchone()
    return row[0]

def load_company(co):
    filepath = os.path.join(RAW_DIR, co["file"] + ".xlsx")
    if not os.path.exists(filepath):
        print(f"  ❌ File not found: {co['file']}.xlsx")
        return False

    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"  ❌ Cannot open: {e}")
        return False

    if 'Data Sheet' not in wb.sheetnames:
        print(f"  ❌ No Data Sheet tab")
        return False

    data = load_data_sheet(wb)

    with engine.connect() as conn:
        sector_id = get_or_create_sector(conn, co["sector"], co["industry"])

        conn.execute(text("""
            INSERT IGNORE INTO companies
                (bse_code, nse_symbol, company_name, sector_id, market_cap)
            VALUES (:bse, :nse, :name, :sid, :mcap)
        """), {"bse": co["bse"], "nse": co["nse"],
               "name": co["name"], "sid": sector_id, "mcap": co["mcap"]})
        conn.commit()

        result = conn.execute(text(
            "SELECT company_id FROM companies WHERE bse_code = :b"
        ), {"b": co["bse"]}).fetchone()
        company_id = result[0]
        print(f"  ✅ ID:{company_id} — {co['name']}")

        years = [excel_date_to_year(y) for y in data.get("PL_Report Date", [])]
        print(f"  📅 {[y for y in years if y]}")

        sales      = data.get("PL_Sales", [])
        raw_mat    = data.get("PL_Raw Material Cost", [None]*12)
        emp_cost   = data.get("PL_Employee Cost", [None]*12)
        oth_inc    = data.get("PL_Other Income", [None]*12)
        dep        = data.get("PL_Depreciation", [None]*12)
        interest   = data.get("PL_Interest", [None]*12)
        pbt        = data.get("PL_Profit before tax", [None]*12)
        tax        = data.get("PL_Tax", [None]*12)
        net_profit = data.get("PL_Net profit", [None]*12)
        dividend   = data.get("PL_Dividend Amount", [None]*12)

        for i, fy in enumerate(years):
            if not fy or i >= len(sales): continue
            try:
                conn.execute(text("""
                    INSERT INTO profit_loss
                        (company_id, fiscal_year, sales, raw_material,
                         employee_cost, other_income, depreciation,
                         interest, profit_before_tax, tax,
                         net_profit, dividend_amount)
                    VALUES (:cid,:fy,:s,:rm,:emp,:oi,:dep,:int,:pbt,:tax,:np,:div)
                    ON DUPLICATE KEY UPDATE sales=VALUES(sales), net_profit=VALUES(net_profit)
                """), {
                    "cid": company_id, "fy": fy,
                    "s":   sales[i]      if i < len(sales)      else None,
                    "rm":  raw_mat[i]    if i < len(raw_mat)    else None,
                    "emp": emp_cost[i]   if i < len(emp_cost)   else None,
                    "oi":  oth_inc[i]    if i < len(oth_inc)    else None,
                    "dep": dep[i]        if i < len(dep)        else None,
                    "int": interest[i]   if i < len(interest)   else None,
                    "pbt": pbt[i]        if i < len(pbt)        else None,
                    "tax": tax[i]        if i < len(tax)        else None,
                    "np":  net_profit[i] if i < len(net_profit) else None,
                    "div": dividend[i]   if i < len(dividend)   else None,
                })
            except Exception as e:
                print(f"    ⚠️ P&L {fy}: {e}")

        bs_years = [excel_date_to_year(y) for y in data.get("BS_Report Date", [])]
        eq_cap   = data.get("BS_Equity Share Capital", [])
        reserves = data.get("BS_Reserves", [])
        borrow   = data.get("BS_Borrowings", [])
        oth_lib  = data.get("BS_Other Liabilities", [])
        tot_lib  = data.get("BS_Total", [])
        net_blk  = data.get("BS_Net Block", [])
        cwip     = data.get("BS_Capital Work in Progress", [])
        invest   = data.get("BS_Investments", [])
        oth_ast  = data.get("BS_Other Assets", [])
        recv     = data.get("BS_Receivables", [])
        inv      = data.get("BS_Inventory", [])
        cash     = data.get("BS_Cash & Bank", [])

        for i, fy in enumerate(bs_years):
            if not fy: continue
            try:
                conn.execute(text("""
                    INSERT INTO balance_sheet
                        (company_id, fiscal_year, equity_capital, reserves,
                         borrowings, other_liabilities, total_liabilities,
                         net_block, cwip, investments, other_assets,
                         receivables, inventory, cash_and_bank)
                    VALUES (:cid,:fy,:eq,:res,:bor,:olib,:tlib,:nb,:cwip,:inv,:oast,:recv,:invt,:cash)
                    ON DUPLICATE KEY UPDATE reserves=VALUES(reserves), borrowings=VALUES(borrowings)
                """), {
                    "cid":  company_id, "fy": fy,
                    "eq":   eq_cap[i]  if i < len(eq_cap)  else None,
                    "res":  reserves[i]if i < len(reserves) else None,
                    "bor":  borrow[i]  if i < len(borrow)  else None,
                    "olib": oth_lib[i] if i < len(oth_lib) else None,
                    "tlib": tot_lib[i] if i < len(tot_lib) else None,
                    "nb":   net_blk[i] if i < len(net_blk) else None,
                    "cwip": cwip[i]    if i < len(cwip)    else None,
                    "inv":  invest[i]  if i < len(invest)  else None,
                    "oast": oth_ast[i] if i < len(oth_ast) else None,
                    "recv": recv[i]    if i < len(recv)    else None,
                    "invt": inv[i]     if i < len(inv)     else None,
                    "cash": cash[i]    if i < len(cash)    else None,
                })
            except Exception as e:
                print(f"    ⚠️ BS {fy}: {e}")

        cf_years = [excel_date_to_year(y) for y in data.get("CF_Report Date", [])]
        op_cf  = data.get("CF_Cash from Operating Activity", [])
        inv_cf = data.get("CF_Cash from Investing Activity", [])
        fin_cf = data.get("CF_Cash from Financing Activity", [])
        net_cf = data.get("CF_Net Cash Flow", [])

        for i, fy in enumerate(cf_years):
            if not fy: continue
            try:
                conn.execute(text("""
                    INSERT INTO cash_flow
                        (company_id, fiscal_year, operating_cf,
                         investing_cf, financing_cf, net_cash_flow)
                    VALUES (:cid,:fy,:ocf,:icf,:fcf,:ncf)
                    ON DUPLICATE KEY UPDATE operating_cf=VALUES(operating_cf)
                """), {
                    "cid": company_id, "fy": fy,
                    "ocf": op_cf[i]  if i < len(op_cf)  else None,
                    "icf": inv_cf[i] if i < len(inv_cf) else None,
                    "fcf": fin_cf[i] if i < len(fin_cf) else None,
                    "ncf": net_cf[i] if i < len(net_cf) else None,
                })
            except Exception as e:
                print(f"    ⚠️ CF {fy}: {e}")

        conn.commit()
        print(f"  ✅ Done: {co['name']}")
        return True

if __name__ == "__main__":
    print("📊 Loading 18 new Nifty 50 companies")
    print("=" * 50)
    loaded = 0
    for co in NEW_COMPANIES:
        print(f"\nLoading: {co['name']}")
        if load_company(co):
            loaded += 1
    print(f"\n{'='*50}")
    print(f"✅ Done! {loaded}/18 companies loaded")