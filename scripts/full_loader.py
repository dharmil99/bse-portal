import os
import sys
sys.path.insert(0, r'C:\Users\Jignesh\Desktop\bse_portal')

import pandas as pd
from openpyxl import load_workbook
from sqlalchemy import text
from scripts.db_connect import get_engine

engine = get_engine()

DATA_FOLDER = r'C:\Users\Jignesh\Desktop\bse_portal\data\raw'

FILE_MAP = {
    "BajajAuto":        "Bajaj Auto Limited",
    "Britannia":        "Britannia Industries",
    "Cipla":            "Cipla Limited",
    "DivisLab":         "Divis Laboratories",
    "DrReddys":         "Dr. Reddys Laboratories",
    "HCLTech":          "HCL Technologies Limited",
    "HDFCBank":         "HDFC Bank Limited",
    "HUL":              "Hindustan Unilever Limited",
    "ICICIBank":        "ICICI Bank Limited",
    "Infosys":          "Infosys Limited",
    "ITC":              "ITC Limited",
    "KotakBank":        "Kotak Mahindra Bank",
    "MahindraMahindra": "Mahindra and Mahindra",
    "Maruti":           "Maruti Suzuki India",
    "NestleIndia":      "Nestle India Limited",
    "SBI":              "State Bank of India",
    "SunPharma":        "Sun Pharmaceutical Industries",
    "TataMotors":       "Tata Motors Limited",
    "TCS":              "Tata Consultancy Services",
    "TechMahindra":     "Tech Mahindra Limited",
    "Wipro":            "Wipro Limited",
    "Hero Motocorp":        "Hero MotoCorp Limited",
    "TVS Motor Co":         "TVS Motor Company Limited",
    "Eicher Motors":        "Eicher Motors Limited",
    "Ashok Leyland":        "Ashok Leyland Limited",
    "Bosch":                "Bosch Limited",
    "MRF":                  "MRF Limited",
    "Apollo Tyres":         "Apollo Tyres Limited",
    "CEAT":                 "CEAT Limited",
    "Balkrishna Inds":      "Balkrishna Industries",
    "Samvardh. Mothe":      "Samvardhana Motherson International Limited",
    "Minda Corp":           "Minda Corporation Limited",
    "Sona BLW Precis":      "Sona BLW Precision Forgings Limited",
    "Endurance Tech":       "Endurance Technologies Limited",
    "Escorts Kubota":       "Escorts Kubota Limited",
    "Force Motors":         "Force Motors Limited",
    "Atul Auto":            "Atul Auto Limited",
    # Banking (incl. NBFC / Insurance)
    "AxisBank":           "Axis Bank Limited",
    "BajajFinance":       "Bajaj Finance Limited",
    "BajajFinserv":       "Bajaj Finserv Limited",
    "ShriramFinance":     "Shriram Finance Limited",
    "HDFCLife":           "HDFC Life Insurance Company Limited",
    "SBILife":            "SBI Life Insurance Company Limited",
    "Jio Financial":      "Jio Financial Services Limited",
 
    # FMCG
    "Trent":              "Trent Limited",
    "Titan Company":      "Titan Company Limited",
    "TataConsumer":       "Tata Consumer Products Limited",
 
    # Pharmaceuticals (incl. hospital chains)
    "ApolloHospitals":    "Apollo Hospitals Enterprise Limited",
    "MaxHealthcare":      "Max Healthcare Institute Limited",
 
    # Energy
    "Coal India":         "Coal India Limited",
    "O N G C":            "Oil and Natural Gas Corporation Limited",
    "NTPC":               "NTPC Limited",
    "Power Grid Corpn":   "Power Grid Corporation of India Limited",
    "Reliance Industries":"Reliance Industries Limited",
 
    # Telecom
    "Bharti Airtel":      "Bharti Airtel Limited",
    "Indus Towers":       "Indus Towers Limited",
 
    # Metals & Mining
    "Adani Enterp":       "Adani Enterprises Limited",
    "JSW Steel":          "JSW Steel Limited",
    "Tata Steel":         "Tata Steel Limited",
    "Hindalco Inds":      "Hindalco Industries Limited",
 
    # Cement
    "UltraTech Cem":      "UltraTech Cement Limited",
    "Grasim Inds":        "Grasim Industries Limited",
 
    # Consumer
    "Eternal":            "Eternal Limited",
 
    # Infrastructure
    "Larsen & Toubro":    "Larsen & Toubro Limited",
}

def excel_date_to_year(val):
    try:
        from datetime import datetime, timedelta
        if isinstance(val, float) or isinstance(val, int):
            dt = datetime(1899, 12, 30) + timedelta(days=int(val))
        else:
            dt = val
        fy = dt.year + 1 if dt.month >= 4 else dt.year
        return f"FY{str(fy)[2:]}"
    except:
        return None

def load_data_sheet(wb):
    ws = wb['Data Sheet']
    data = {}
    current_section = None

    for row in ws.iter_rows(values_only=True):
        if not any(row):
            continue
        label = str(row[0]).strip() if row[0] else ""

        if label == "PROFIT & LOSS":
            current_section = "PL"
            continue
        elif label == "BALANCE SHEET":
            current_section = "BS"
            continue
        elif label == "CASH FLOW:":
            current_section = "CF"
            continue
        elif label == "Quarters":
            current_section = "Q"
            continue
        elif label == "META":
            current_section = "META"
            continue

        values = [v for v in row[1:] if v is not None]
        if values:
            data[f"{current_section}_{label}"] = values

    return data

def load_company(filepath):
    filename = os.path.basename(filepath)
    file_key = filename.replace('.xlsx', '').strip()

    if file_key.startswith('~$'):
        return

    print(f"\n📂 Loading: {filename}")

    mapped_name = FILE_MAP.get(file_key)
    if not mapped_name:
        print(f"  ⚠️ No mapping found for: {file_key}")
        return

    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"  ❌ Could not open: {e}")
        return

    if 'Data Sheet' not in wb.sheetnames:
        print(f"  ❌ No Data Sheet tab found")
        return

    data = load_data_sheet(wb)

    with engine.connect() as conn:
        result = conn.execute(text(
            "SELECT company_id FROM companies WHERE company_name = :name"
        ), {"name": mapped_name}).fetchone()

        if not result:
            print(f"  ⚠️ Company not found in DB: {mapped_name}")
            return

        company_id = result[0]
        print(f"  ✅ Matched: {mapped_name} (id={company_id})")

        # --- P&L ---
        years_raw     = data.get("PL_Report Date", [])
        years         = [excel_date_to_year(y) for y in years_raw]
        sales         = data.get("PL_Sales", [])
        raw_mat       = data.get("PL_Raw Material Cost", [None]*10)
        emp_cost      = data.get("PL_Employee Cost", [None]*10)
        other_income  = data.get("PL_Other Income", [None]*10)
        depreciation  = data.get("PL_Depreciation", [None]*10)
        interest      = data.get("PL_Interest", [None]*10)
        pbt           = data.get("PL_Profit before tax", [None]*10)
        tax           = data.get("PL_Tax", [None]*10)
        net_profit    = data.get("PL_Net profit", [None]*10)
        dividend      = data.get("PL_Dividend Amount", [None]*10)

        print(f"  📅 Years: {years}")

        for i, fy in enumerate(years):
            if not fy or i >= len(sales):
                continue
            try:
                conn.execute(text("""
                    INSERT INTO profit_loss
                        (company_id, fiscal_year, sales, raw_material, employee_cost,
                         other_income, depreciation, interest,
                         profit_before_tax, tax, net_profit, dividend_amount)
                    VALUES
                        (:cid, :fy, :sales, :rm, :emp, :oi, :dep, :int, :pbt, :tax, :np, :div)
                    ON DUPLICATE KEY UPDATE
                        sales=VALUES(sales), net_profit=VALUES(net_profit)
                """), {
                    "cid":   company_id, "fy": fy,
                    "sales": sales[i] if i < len(sales) else None,
                    "rm":    raw_mat[i] if i < len(raw_mat) else None,
                    "emp":   emp_cost[i] if i < len(emp_cost) else None,
                    "oi":    other_income[i] if i < len(other_income) else None,
                    "dep":   depreciation[i] if i < len(depreciation) else None,
                    "int":   interest[i] if i < len(interest) else None,
                    "pbt":   pbt[i] if i < len(pbt) else None,
                    "tax":   tax[i] if i < len(tax) else None,
                    "np":    net_profit[i] if i < len(net_profit) else None,
                    "div":   dividend[i] if i < len(dividend) else None,
                })
            except Exception as e:
                print(f"  ⚠️ P&L error {fy}: {e}")

        # --- Balance Sheet ---
        bs_years  = [excel_date_to_year(y) for y in data.get("BS_Report Date", [])]
        eq_cap    = data.get("BS_Equity Share Capital", [])
        reserves  = data.get("BS_Reserves", [])
        borrow    = data.get("BS_Borrowings", [])
        other_lib = data.get("BS_Other Liabilities", [])
        total_lib = data.get("BS_Total", [])
        net_block = data.get("BS_Net Block", [])
        cwip      = data.get("BS_Capital Work in Progress", [])
        invest    = data.get("BS_Investments", [])
        other_ast = data.get("BS_Other Assets", [])
        recv      = data.get("BS_Receivables", [])
        inv       = data.get("BS_Inventory", [])
        cash      = data.get("BS_Cash & Bank", [])

        for i, fy in enumerate(bs_years):
            if not fy:
                continue
            try:
                conn.execute(text("""
                    INSERT INTO balance_sheet
                        (company_id, fiscal_year, equity_capital, reserves,
                         borrowings, other_liabilities, total_liabilities,
                         net_block, cwip, investments, other_assets,
                         receivables, inventory, cash_and_bank)
                    VALUES
                        (:cid, :fy, :eq, :res, :bor, :olib, :tlib,
                         :nb, :cwip, :inv, :oast, :recv, :invt, :cash)
                    ON DUPLICATE KEY UPDATE
                        reserves=VALUES(reserves), borrowings=VALUES(borrowings)
                """), {
                    "cid":  company_id, "fy": fy,
                    "eq":   eq_cap[i] if i < len(eq_cap) else None,
                    "res":  reserves[i] if i < len(reserves) else None,
                    "bor":  borrow[i] if i < len(borrow) else None,
                    "olib": other_lib[i] if i < len(other_lib) else None,
                    "tlib": total_lib[i] if i < len(total_lib) else None,
                    "nb":   net_block[i] if i < len(net_block) else None,
                    "cwip": cwip[i] if i < len(cwip) else None,
                    "inv":  invest[i] if i < len(invest) else None,
                    "oast": other_ast[i] if i < len(other_ast) else None,
                    "recv": recv[i] if i < len(recv) else None,
                    "invt": inv[i] if i < len(inv) else None,
                    "cash": cash[i] if i < len(cash) else None,
                })
            except Exception as e:
                print(f"  ⚠️ BS error {fy}: {e}")

        # --- Cash Flow ---
        cf_years = [excel_date_to_year(y) for y in data.get("CF_Report Date", [])]
        op_cf    = data.get("CF_Cash from Operating Activity", [])
        inv_cf   = data.get("CF_Cash from Investing Activity", [])
        fin_cf   = data.get("CF_Cash from Financing Activity", [])
        net_cf   = data.get("CF_Net Cash Flow", [])

        for i, fy in enumerate(cf_years):
            if not fy:
                continue
            try:
                conn.execute(text("""
                    INSERT INTO cash_flow
                        (company_id, fiscal_year, operating_cf,
                         investing_cf, financing_cf, net_cash_flow)
                    VALUES (:cid, :fy, :ocf, :icf, :fcf, :ncf)
                    ON DUPLICATE KEY UPDATE
                        operating_cf=VALUES(operating_cf)
                """), {
                    "cid": company_id, "fy": fy,
                    "ocf": op_cf[i] if i < len(op_cf) else None,
                    "icf": inv_cf[i] if i < len(inv_cf) else None,
                    "fcf": fin_cf[i] if i < len(fin_cf) else None,
                    "ncf": net_cf[i] if i < len(net_cf) else None,
                })
            except Exception as e:
                print(f"  ⚠️ CF error {fy}: {e}")

        conn.commit()
        print(f"  ✅ All data loaded for {mapped_name}")

if __name__ == "__main__":
    files = [f for f in os.listdir(DATA_FOLDER) if f.endswith('.xlsx')]
    print(f"Found {len(files)} Excel files")
    for f in sorted(files):
        load_company(os.path.join(DATA_FOLDER, f))
    print("\n🎉 Full load complete!")