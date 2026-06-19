import os
import sys
sys.path.insert(0, r'C:\Users\Jignesh\Desktop\bse_portal')

import pandas as pd
from sqlalchemy import text
from scripts.db_connect import get_engine
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

engine = get_engine()

OUTPUT_PATH = r'C:\Users\Jignesh\Desktop\bse_portal\BenchmarkIQ_Excellence_Model_Filled.xlsx'
TEMPLATE_PATH = r'C:\Users\Jignesh\Desktop\BenchmarkIQ_Excellence_Model.xlsx'

AUTO_COMPANIES = [
    "Tata Motors Limited", "Maruti Suzuki India", "Mahindra and Mahindra",
    "Bajaj Auto Limited", "Hero MotoCorp Limited", "TVS Motor Company",
    "Eicher Motors Limited", "Ashok Leyland Limited", "Bosch Limited",
    "MRF Limited", "Apollo Tyres Limited", "CEAT Limited",
    "Balkrishna Industries", "Samvardhana Motherson", "Minda Industries Limited",
    "Sona BLW Precision", "Endurance Technologies", "Escorts Kubota Limited",
    "Force Motors Limited", "Atul Auto Limited"
]

RATIO_WEIGHTS = {
    "Net Profit Margin":       0.10,
    "EBITDA Margin":           0.10,
    "ROE":                     0.08,
    "ROCE":                    0.07,
    "Operating Profit Margin": 0.05,
    "Revenue Growth YoY":      0.10,
    "3Y Revenue CAGR":         0.08,
    "NP Growth YoY":           0.07,
    "Asset Turnover":          0.07,
    "Debtor Days":             0.05,
    "Inventory Turnover":      0.08,
    "Debt to Equity":          0.08,
    "Interest Coverage":       0.07,
    "EPS Growth YoY":          0.05,
    "Current Ratio":           0.05,
}

HIGHER_BETTER = {
    "Net Profit Margin": True, "EBITDA Margin": True,
    "ROE": True, "ROCE": True, "Operating Profit Margin": True,
    "Revenue Growth YoY": True, "3Y Revenue CAGR": True,
    "NP Growth YoY": True, "Asset Turnover": True,
    "Debtor Days": False, "Inventory Turnover": True,
    "Debt to Equity": False, "Interest Coverage": True,
    "EPS Growth YoY": True, "Current Ratio": True,
}

def fetch_company_data(company_name):
    with engine.connect() as conn:
        # Get company_id
        result = conn.execute(text(
            "SELECT company_id FROM companies WHERE company_name = :name"
        ), {"name": company_name}).fetchone()
        if not result:
            return None, None, None
        cid = result[0]

        pl = pd.read_sql(text("""
            SELECT fiscal_year, sales, net_profit, raw_material,
                   employee_cost, depreciation, interest,
                   profit_before_tax, other_income
            FROM profit_loss WHERE company_id = :cid
            ORDER BY fiscal_year
        """), conn, params={"cid": cid})

        bs = pd.read_sql(text("""
            SELECT fiscal_year, equity_capital, reserves, borrowings,
                   total_assets, net_block, receivables, inventory,
                   cash_and_bank, other_liabilities
            FROM balance_sheet WHERE company_id = :cid
            ORDER BY fiscal_year
        """), conn, params={"cid": cid})

        fr = pd.read_sql(text("""
            SELECT quarter, roe, roce, debt_to_equity, net_margin,
                   ebitda_margin, revenue_growth, cagr_3y
            FROM financial_ratios WHERE company_id = :cid
            ORDER BY quarter DESC LIMIT 1
        """), conn, params={"cid": cid})

    return pl, bs, fr

def calculate_ratios(pl, bs, fr):
    if pl is None or pl.empty:
        return {}

    ratios = {}

    # Merge on fiscal_year
    merged = pd.merge(pl, bs, on="fiscal_year", how="inner")
    if merged.empty:
        return {}

    latest = merged.iloc[-1]
    prev   = merged.iloc[-2] if len(merged) > 1 else latest
    old3   = merged.iloc[-4] if len(merged) > 3 else merged.iloc[0]

    def safe_div(a, b, mult=1):
        try:
            if b and b != 0:
                return round(float(a) / float(b) * mult, 2)
        except:
            pass
        return None

    # Profitability
    ratios["Net Profit Margin"]       = safe_div(latest["net_profit"], latest["sales"], 100)
    ebitda = (latest["net_profit"] or 0) + (latest["interest"] or 0) + (latest["depreciation"] or 0)
    ratios["EBITDA Margin"]           = safe_div(ebitda, latest["sales"], 100)
    equity = (latest.get("equity_capital") or 0) + (latest.get("reserves") or 0)
    ratios["ROE"]                     = safe_div(latest["net_profit"], equity, 100)
    cap_emp = (latest.get("total_assets") or 0) - (latest.get("other_liabilities") or 0)
    ebit = (latest["net_profit"] or 0) + (latest["interest"] or 0)
    ratios["ROCE"]                    = safe_div(ebit, cap_emp, 100)
    op_profit = (latest["net_profit"] or 0) + (latest["interest"] or 0) + (latest["depreciation"] or 0)
    ratios["Operating Profit Margin"] = safe_div(op_profit, latest["sales"], 100)

    # Growth
    ratios["Revenue Growth YoY"]      = safe_div(latest["sales"] - prev["sales"], prev["sales"], 100)
    ratios["3Y Revenue CAGR"]         = round((float(latest["sales"]) / float(old3["sales"])) ** (1/3) - 1, 4) * 100 if old3["sales"] and old3["sales"] != 0 else None
    ratios["NP Growth YoY"]           = safe_div(latest["net_profit"] - prev["net_profit"], prev["net_profit"], 100)
    ratios["EPS Growth YoY"]          = ratios["NP Growth YoY"]  # proxy

    # Efficiency
    ratios["Asset Turnover"]          = safe_div(latest["sales"], latest.get("total_assets"))
    ratios["Debtor Days"]             = safe_div((latest.get("receivables") or 0), latest["sales"], 365)
    ratios["Inventory Turnover"]      = safe_div(latest["sales"], latest.get("inventory"))

    # Safety
    ratios["Debt to Equity"]          = safe_div(latest.get("borrowings"), equity)
    ratios["Interest Coverage"]       = safe_div(ebit, latest["interest"])
    curr_assets = (latest.get("receivables") or 0) + (latest.get("inventory") or 0) + (latest.get("cash_and_bank") or 0)
    ratios["Current Ratio"]           = safe_div(curr_assets, latest.get("other_liabilities"))

    return ratios

def percentile_rank(value, all_values, higher_better=True):
    """Rank 0-100 based on percentile within group"""
    valid = [v for v in all_values if v is not None]
    if not valid or value is None:
        return 50  # neutral if no data
    if higher_better:
        rank = sum(1 for v in valid if v <= value) / len(valid) * 100
    else:
        rank = sum(1 for v in valid if v >= value) / len(valid) * 100
    return round(rank, 1)

def main():
    print("📊 BenchmarkIQ Excellence Model Calculator")
    print("=" * 50)

    # Fetch all company ratios
    all_ratios = {}
    for company in AUTO_COMPANIES:
        print(f"  Fetching: {company}")
        pl, bs, fr = fetch_company_data(company)
        ratios = calculate_ratios(pl, bs, fr)
        all_ratios[company] = ratios

    print("\n✅ All ratios calculated")

    # Calculate percentile scores
    ratio_names = list(RATIO_WEIGHTS.keys())
    percentile_scores = {}

    for ratio in ratio_names:
        all_vals = [all_ratios[c].get(ratio) for c in AUTO_COMPANIES]
        for company in AUTO_COMPANIES:
            val = all_ratios[company].get(ratio)
            score = percentile_rank(val, all_vals, HIGHER_BETTER[ratio])
            if company not in percentile_scores:
                percentile_scores[company] = {}
            percentile_scores[company][ratio] = score

    # Calculate weighted total scores
    total_scores = {}
    for company in AUTO_COMPANIES:
        score = sum(
            percentile_scores[company].get(r, 50) * w
            for r, w in RATIO_WEIGHTS.items()
        )
        total_scores[company] = round(score, 1)

    # Sort by score
    ranked = sorted(total_scores.items(), key=lambda x: x[1], reverse=True)

    print("\n🏆 EXCELLENCE RANKINGS — AUTOMOBILE SECTOR")
    print("=" * 50)
    for i, (company, score) in enumerate(ranked):
        if score >= 85:   tier = "🏆 Excellence Leader"
        elif score >= 70: tier = "🥇 High Performer"
        elif score >= 55: tier = "🟢 Above Average"
        elif score >= 40: tier = "🟡 Average"
        elif score >= 25: tier = "🟠 Below Average"
        else:             tier = "🔴 Needs Improvement"
        print(f"  {i+1:2}. {company:<30} {score:5.1f}  {tier}")

    # Build output DataFrame
    rows = []
    for rank, (company, score) in enumerate(ranked):
        row = {"Rank": rank+1, "Company": company, "Excellence Score": score}
        for r in ratio_names:
            row[r] = round(all_ratios[company].get(r) or 0, 2)
        rows.append(row)

    df = pd.DataFrame(rows)
    df.to_excel(OUTPUT_PATH, index=False, sheet_name="Excellence Rankings")
    print(f"\n✅ Saved to: {OUTPUT_PATH}")

if __name__ == "__main__":
    main()