# generate_ceo_presentation_combined.py
#
# Reads the 62-company, 10-sector dataset from
# BenchmarkIQ_Excellence_Model_Filled.xlsx ("All Sectors" sheet) and
# generates ONE combined workbook covering every sector and company,
# using the same navy/gold formatting as the per-sector presentations.
#
# Run from the scripts folder:  python generate_ceo_presentation_combined.py
# Input expected at the project root:  BenchmarkIQ_Excellence_Model_Filled.xlsx
# Output written to the project root:  CEO_Excellence_Presentation_AllSectors.xlsx

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

INPUT_PATH  = r"C:\Users\Jignesh\Desktop\bse_portal\BenchmarkIQ_Excellence_Model_Filled.xlsx"
OUTPUT_PATH = r"C:\Users\Jignesh\Desktop\bse_portal\CEO_Excellence_Presentation_AllSectors.xlsx"

# ── Colors (sampled from the existing template) ──────────────────────────
NAVY        = "1F3864"
GOLD        = "C9A84C"
LIGHT_BLUE  = "D9E1F2"
WHITE       = "FFFFFF"
DARK_GREY   = "404040"
DARK_GREEN  = "375623"
LIGHT_GREY  = "F2F2F2"
HEAT_GREEN  = "C6EFCE"
HEAT_YELLOW = "FFEB9C"
HEAT_RED    = "FFC7CE"

RATIO_WEIGHTS = {
    "Net Profit Margin":       0.10,
    "EBITDA Margin":           0.10,
    "ROE":                     0.08,
    "ROCE":                    0.05,
    "Operating Profit Margin": 0.02,
    "Revenue Growth YoY":      0.10,
    "3Y Revenue CAGR":         0.08,
    "NP Growth YoY":           0.05,
    "EPS Growth YoY":          0.02,
    "Asset Turnover":          0.07,
    "Debtor Days":             0.05,
    "Inventory Turnover":      0.08,
    "Debt to Equity":          0.08,
    "Interest Coverage":       0.07,
    "Current Ratio":           0.05,
}

HIGHER_BETTER = {
    "Net Profit Margin": True, "EBITDA Margin": True,
    "ROE": True, "ROCE": True, "Operating Profit Margin": True,
    "Revenue Growth YoY": True, "3Y Revenue CAGR": True,
    "NP Growth YoY": True, "EPS Growth YoY": True,
    "Asset Turnover": True, "Debtor Days": False,
    "Inventory Turnover": True, "Debt to Equity": False,
    "Interest Coverage": True, "Current Ratio": True,
}

CATEGORIES = {
    "Profitability": ["Net Profit Margin", "EBITDA Margin", "ROE", "ROCE", "Operating Profit Margin"],
    "Growth":        ["Revenue Growth YoY", "3Y Revenue CAGR", "NP Growth YoY", "EPS Growth YoY"],
    "Efficiency":    ["Asset Turnover", "Debtor Days", "Inventory Turnover"],
    "Safety":        ["Debt to Equity", "Interest Coverage", "Current Ratio"],
}

RATIO_NAMES = list(RATIO_WEIGHTS.keys())


def tier_label(score):
    if score >= 85:   return "Excellence Leader"
    elif score >= 70: return "High Performer"
    elif score >= 55: return "Above Average"
    elif score >= 40: return "Average"
    elif score >= 25: return "Below Average"
    else:             return "Needs Improvement"


def tier_fill_color(tier):
    return {
        "Excellence Leader":  DARK_GREEN,
        "High Performer":     DARK_GREEN,
        "Above Average":      "538135",
        "Average":            "BF8F00",
        "Below Average":      "C45911",
        "Needs Improvement":  "C00000",
    }.get(tier, DARK_GREY)


def percentile_rank(value, all_values, higher_better=True):
    valid = [v for v in all_values if v is not None]
    if not valid or value is None:
        return 50.0
    if higher_better:
        rank = sum(1 for v in valid if v <= value) / len(valid) * 100
    else:
        rank = sum(1 for v in valid if v >= value) / len(valid) * 100
    return round(rank, 1)


def short_name(company):
    name = company.strip()
    for suffix in [" Limited", " Ltd.", " Ltd"]:
        if name.endswith(suffix):
            name = name[: -len(suffix)]
            break
    return name.strip()


def load_all_companies():
    """Read the All Sectors sheet. Returns (companies list, sectors dict)."""
    wb = openpyxl.load_workbook(INPUT_PATH, data_only=True)
    ws = wb["All Sectors"]
    headers = [c.value for c in ws[1]]

    companies = []
    sectors = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        record = dict(zip(headers, row))
        companies.append(record)
        sectors.setdefault(record["Sector"], []).append(record)

    return companies, sectors


def compute_category_scores_within_sector(companies):
    """
    Percentiles and category scores computed WITHIN each company's own
    sector (consistent with how the original Excellence Score was derived).
    Returns dicts keyed by company name.
    """
    percentile_scores = {}
    category_scores = {}

    by_sector = {}
    for c in companies:
        by_sector.setdefault(c["Sector"], []).append(c)

    for sector, sector_companies in by_sector.items():
        for ratio in RATIO_NAMES:
            all_vals = [c.get(ratio) for c in sector_companies]
            for c in sector_companies:
                name = c["Company"]
                pct = percentile_rank(c.get(ratio), all_vals, HIGHER_BETTER[ratio])
                percentile_scores.setdefault(name, {})[ratio] = pct

        for c in sector_companies:
            name = c["Company"]
            category_scores[name] = {}
            for cat, ratios in CATEGORIES.items():
                pct_values = [percentile_scores[name][r] for r in ratios]
                category_scores[name][cat] = round(sum(pct_values) / len(pct_values), 1)

    return percentile_scores, category_scores


def heat_color(pct):
    if pct is None:
        return LIGHT_GREY
    if pct >= 75:
        return HEAT_GREEN
    elif pct >= 50:
        return HEAT_YELLOW
    else:
        return HEAT_RED


def build_cover(wb, companies, sectors):
    ws = wb.create_sheet("Cover")
    for col in "ABCDEFGHIJKLMN":
        ws.column_dimensions[col].width = 10

    for r in range(1, 40):
        for c in range(1, 15):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=NAVY)

    ws["C9"] = "BenchmarkIQ"
    ws["C9"].font = Font(bold=True, size=36, color=WHITE)
    ws["C11"] = "ALL SECTORS — EXCELLENCE MODEL"
    ws["C11"].font = Font(bold=True, size=20, color=GOLD)
    ws["C13"] = "Excellence Model — CEO Presentation"
    ws["C13"].font = Font(size=16, color="BDD7EE")

    overall_top = max(companies, key=lambda c: c["Excellence Score"])
    ws["C16"] = (f"Top overall (sector-relative): {short_name(overall_top['Company'])} "
                 f"({overall_top['Sector']}) — {overall_top['Excellence Score']:.1f} / 100")
    ws["C16"].font = Font(bold=True, size=12, color=WHITE)
    ws["C18"] = f"{len(companies)} companies  ·  {len(sectors)} sectors  ·  15 financial ratios per company"
    ws["C18"].font = Font(size=11, color="BDD7EE")
    ws["C20"] = "Note: scores are percentile-ranked within each company's own sector,"
    ws["C20"].font = Font(size=9, italic=True, color="BDD7EE")
    ws["C21"] = "so cross-sector comparisons reflect relative standing, not absolute equivalence."
    ws["C21"].font = Font(size=9, italic=True, color="BDD7EE")


def build_sector_summary(wb, sectors, category_scores):
    ws = wb.create_sheet("Sector Summary")
    ws["A2"] = "Sector Comparison — Overview"
    ws["A2"].font = Font(bold=True, size=14, color=WHITE)
    ws["A2"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells("A2:F2")

    headers = ["Sector", "Companies", "Avg Score", "Top Performer", "Top Score", "Top Tier"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = Font(bold=True, color=NAVY)
        cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)

    sector_stats = []
    for sector, comps in sectors.items():
        avg_score = sum(c["Excellence Score"] for c in comps) / len(comps)
        top = max(comps, key=lambda c: c["Excellence Score"])
        sector_stats.append((sector, len(comps), round(avg_score, 1), top["Company"],
                              top["Excellence Score"], tier_label(top["Excellence Score"])))

    sector_stats.sort(key=lambda x: -x[2])

    for i, (sector, n, avg, top_co, top_score, top_tier) in enumerate(sector_stats):
        r = 5 + i
        ws.cell(row=r, column=1, value=sector).font = Font(color=DARK_GREY)
        ws.cell(row=r, column=2, value=n).font = Font(color=DARK_GREY)
        ws.cell(row=r, column=3, value=avg).font = Font(bold=True, color=DARK_GREEN)
        ws.cell(row=r, column=4, value=short_name(top_co)).font = Font(color=DARK_GREY)
        ws.cell(row=r, column=5, value=top_score).font = Font(bold=True, color=GOLD)
        tier_cell = ws.cell(row=r, column=6, value=top_tier)
        tier_cell.font = Font(bold=True, color=WHITE)
        tier_cell.fill = PatternFill("solid", fgColor=tier_fill_color(top_tier))

    widths = [22, 12, 12, 28, 11, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_rankings_by_sector(wb, sectors, category_scores):
    ws = wb.create_sheet("Rankings by Sector")
    ws["A2"] = "All Companies — Grouped by Sector, Ranked Within Sector"
    ws["A2"].font = Font(bold=True, size=14, color=WHITE)
    ws["A2"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells("A2:I2")

    headers = ["Sector", "Sector Rank", "Company", "Excellence Score", "Tier",
               "Profitability", "Growth", "Efficiency", "Safety"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = Font(bold=True, color=NAVY)
        cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)

    row_cursor = 5
    for sector in sorted(sectors.keys()):
        comps = sorted(sectors[sector], key=lambda c: -c["Excellence Score"])
        for i, c in enumerate(comps):
            name = c["Company"]
            score = c["Excellence Score"]
            tier = tier_label(score)
            cat = category_scores[name]

            ws.cell(row=row_cursor, column=1, value=sector).font = Font(color=DARK_GREY, size=9)
            ws.cell(row=row_cursor, column=2, value=i + 1).font = Font(bold=True, color=GOLD if i < 3 else DARK_GREY)
            ws.cell(row=row_cursor, column=3, value=short_name(name)).font = Font(color=DARK_GREY)
            ws.cell(row=row_cursor, column=4, value=score).font = Font(bold=True, color=DARK_GREEN)
            tier_cell = ws.cell(row=row_cursor, column=5, value=tier)
            tier_cell.font = Font(bold=True, color=WHITE)
            tier_cell.fill = PatternFill("solid", fgColor=tier_fill_color(tier))
            ws.cell(row=row_cursor, column=6, value=cat["Profitability"]).font = Font(color=DARK_GREY)
            ws.cell(row=row_cursor, column=7, value=cat["Growth"]).font = Font(color=DARK_GREY)
            ws.cell(row=row_cursor, column=8, value=cat["Efficiency"]).font = Font(color=DARK_GREY)
            ws.cell(row=row_cursor, column=9, value=cat["Safety"]).font = Font(color=DARK_GREY)
            row_cursor += 1

    widths = [22, 12, 28, 16, 18, 14, 12, 12, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_overall_ranking(wb, companies, category_scores):
    ws = wb.create_sheet("Overall Ranking (Flat)")
    ws["A2"] = "Overall Ranking — All 62 Companies, Score-Only (Sector-Relative)"
    ws["A2"].font = Font(bold=True, size=14, color=WHITE)
    ws["A2"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells("A2:I2")

    ws["A3"] = ("Note: each score is percentile-ranked within its own sector — "
                "a 70 in a 5-company sector is not strictly equivalent to a 70 in a 20-company sector.")
    ws["A3"].font = Font(size=8, italic=True, color=DARK_GREY)
    ws.merge_cells("A3:I3")

    headers = ["Overall Rank", "Company", "Sector", "Excellence Score", "Tier",
               "Profitability", "Growth", "Efficiency", "Safety"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=c, value=h)
        cell.font = Font(bold=True, color=NAVY)
        cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)

    ranked = sorted(companies, key=lambda c: -c["Excellence Score"])

    for i, c in enumerate(ranked):
        r = 6 + i
        name = c["Company"]
        score = c["Excellence Score"]
        tier = tier_label(score)
        cat = category_scores[name]

        ws.cell(row=r, column=1, value=i + 1).font = Font(bold=True, color=GOLD if i < 10 else DARK_GREY)
        ws.cell(row=r, column=2, value=short_name(name)).font = Font(color=DARK_GREY)
        ws.cell(row=r, column=3, value=c["Sector"]).font = Font(color=DARK_GREY, size=9)
        ws.cell(row=r, column=4, value=score).font = Font(bold=True, color=DARK_GREEN)
        tier_cell = ws.cell(row=r, column=5, value=tier)
        tier_cell.font = Font(bold=True, color=WHITE)
        tier_cell.fill = PatternFill("solid", fgColor=tier_fill_color(tier))
        ws.cell(row=r, column=6, value=cat["Profitability"]).font = Font(color=DARK_GREY)
        ws.cell(row=r, column=7, value=cat["Growth"]).font = Font(color=DARK_GREY)
        ws.cell(row=r, column=8, value=cat["Efficiency"]).font = Font(color=DARK_GREY)
        ws.cell(row=r, column=9, value=cat["Safety"]).font = Font(color=DARK_GREY)

    widths = [12, 28, 22, 16, 18, 14, 12, 12, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_full_heatmap(wb, companies, percentile_scores):
    ws = wb.create_sheet("Full Ratio Heatmap")
    ws["A2"] = "Ratio Heatmap — All 62 Companies, Grouped by Sector"
    ws["A2"].font = Font(bold=True, size=13, color=NAVY)

    short_labels = [r.replace(" ", "\n") for r in RATIO_NAMES]
    headers = ["Sector", "Company", "Score"] + short_labels
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = Font(bold=True, size=9, color=NAVY)
        cell.alignment = Alignment(wrap_text=True, horizontal="center")
        cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)

    for c, ratio in enumerate(RATIO_NAMES, start=4):
        pct_cell = ws.cell(row=5, column=c, value=f"{int(RATIO_WEIGHTS[ratio]*100)}%")
        pct_cell.font = Font(size=8, italic=True, color=DARK_GREY)
        pct_cell.fill = PatternFill("solid", fgColor=LIGHT_GREY)

    sorted_companies = sorted(companies, key=lambda c: (c["Sector"], -c["Excellence Score"]))

    for i, c in enumerate(sorted_companies):
        r = 6 + i
        name = c["Company"]
        ws.cell(row=r, column=1, value=c["Sector"]).font = Font(size=8, color=DARK_GREY)
        ws.cell(row=r, column=2, value=short_name(name)).font = Font(size=9)
        ws.cell(row=r, column=3, value=c["Excellence Score"]).font = Font(bold=True, size=9)

        for col, ratio in enumerate(RATIO_NAMES, start=4):
            val = c.get(ratio)
            pct = percentile_scores[name].get(ratio)
            cell = ws.cell(row=r, column=col, value=round(val, 1) if val is not None else None)
            cell.fill = PatternFill("solid", fgColor=heat_color(pct))
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(size=8)

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 8
    for i in range(4, 4 + len(RATIO_NAMES)):
        ws.column_dimensions[get_column_letter(i)].width = 9
    ws.freeze_panes = "D6"


def build_top_per_sector(wb, sectors, percentile_scores):
    ws = wb.create_sheet("Top Performer Per Sector")
    ws["A2"] = "Top Performer Per Sector — Detailed Ratio Profile"
    ws["A2"].font = Font(bold=True, size=13, color=NAVY)

    sorted_sectors = sorted(sectors.keys())
    block_starts = list(range(1, 1 + len(sorted_sectors) * 4, 4))

    for idx, sector in enumerate(sorted_sectors):
        comps = sectors[sector]
        top = max(comps, key=lambda c: c["Excellence Score"])
        name = top["Company"]
        start_col = block_starts[idx]
        end_col = start_col + 2

        ws.cell(row=4, column=start_col, value=f"{sector}: {short_name(name)}").font = Font(bold=True, color=NAVY, size=9)
        ws.merge_cells(start_row=4, start_column=start_col, end_row=4, end_column=end_col)

        ws.cell(row=5, column=start_col, value=f"Score: {top['Excellence Score']:.1f}").font = Font(italic=True, size=8)
        ws.merge_cells(start_row=5, start_column=start_col, end_row=5, end_column=end_col)

        for c, label in zip([start_col, start_col + 1, start_col + 2], ["Metric", "Value", "Pct"]):
            cell = ws.cell(row=6, column=c, value=label)
            cell.font = Font(bold=True, size=8)
            cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)

        for i, ratio in enumerate(RATIO_NAMES):
            r = 7 + i
            val = top.get(ratio)
            pct = percentile_scores[name].get(ratio)
            ws.cell(row=r, column=start_col, value=ratio).font = Font(size=7)
            ws.cell(row=r, column=start_col + 1,
                    value=round(val, 1) if val is not None else None).font = Font(size=7)
            ws.cell(row=r, column=start_col + 2,
                    value=f"{pct:.0f}th" if pct is not None else "N/A").font = Font(size=7)

    for i in range(1, max(block_starts) + 4):
        ws.column_dimensions[get_column_letter(i)].width = 11


def build_methodology(wb, n_companies, n_sectors):
    ws = wb.create_sheet("Methodology")
    ws["A2"] = "Methodology — How Excellence Scores Are Calculated"
    ws["A2"].font = Font(bold=True, size=13, color=NAVY)

    lines = [
        (4, "OVERVIEW", True),
        (5, f"The Excellence Model ranks {n_companies} Indian companies across {n_sectors} sectors "
            f"using {len(RATIO_NAMES)} financial ratios across 4 categories.", False),
        (6, "Each company is percentile-ranked (0-100) against peers WITHIN ITS OWN SECTOR, "
            "then a weighted average score is computed.", False),
        (7, "Scores are therefore sector-relative: they show standing within a sector, not an "
            "absolute cross-sector ranking.", False),
        (9, "SCORING METHODOLOGY", True),
        (10, "Step 1: Calculate 15 ratios from P&L and Balance Sheet data for each company (latest FY).", False),
        (11, "Step 2: Rank each company's ratio against peers in the SAME SECTOR using percentile ranking.", False),
        (12, "Step 3: Apply category weights and compute a weighted Excellence Score (0-100).", False),
        (14, "TIER CLASSIFICATION", True),
        (15, "85-100  ->  Excellence Leader   |  70-84  ->  High Performer   |  55-69  ->  Above Average", False),
        (16, "40-54   ->  Average              |  25-39  ->  Below Average    |  0-24   ->  Needs Improvement", False),
        (18, "DATA SOURCE", True),
        (19, "BSE/NSE filings via Screener.in  |  FY ending March 2026", False),
        (21, "COLOR CODING (HEATMAP)", True),
        (22, "Green  =  Top quartile within sector (75th percentile+)", False),
        (23, "Yellow =  Middle within sector (50th-75th percentile)", False),
        (24, "Red    =  Bottom half within sector (below 50th percentile)", False),
        (26, "RATIO WEIGHTS", True),
    ]
    for r, text, bold in lines:
        cell = ws.cell(row=r, column=1, value=text)
        cell.font = Font(bold=bold, size=12 if bold and r == 2 else 10, color=NAVY if bold else DARK_GREY)

    cat_map = {r: cat for cat, ratios in CATEGORIES.items() for r in ratios}
    for i, ratio in enumerate(RATIO_NAMES):
        r = 27 + i
        cat = cat_map[ratio]
        pct = int(RATIO_WEIGHTS[ratio] * 100)
        ws.cell(row=r, column=1, value=f"  {ratio}  ({cat})  ->  {pct}%").font = Font(size=9, color=DARK_GREY)

    ws.column_dimensions["A"].width = 100


def main():
    if not os.path.exists(INPUT_PATH):
        print(f"ERROR: Input file not found at {INPUT_PATH}")
        print("Make sure BenchmarkIQ_Excellence_Model_Filled.xlsx is at the project root.")
        return

    companies, sectors = load_all_companies()
    print(f"Loaded {len(companies)} companies across {len(sectors)} sectors")

    percentile_scores, category_scores = compute_category_scores_within_sector(companies)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_cover(wb, companies, sectors)
    build_sector_summary(wb, sectors, category_scores)
    build_rankings_by_sector(wb, sectors, category_scores)
    build_overall_ranking(wb, companies, category_scores)
    build_full_heatmap(wb, companies, percentile_scores)
    build_top_per_sector(wb, sectors, percentile_scores)
    build_methodology(wb, len(companies), len(sectors))

    wb.save(OUTPUT_PATH)
    print(f"Saved combined workbook: {OUTPUT_PATH}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()