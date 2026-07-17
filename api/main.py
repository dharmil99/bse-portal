from fastapi import FastAPI, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy import text
import sys
sys.path.insert(0, r'C:\Users\Jignesh\Desktop\bse_portal')
from scripts.db_connect import get_engine
from app.components.pdf_generator import generate_pdf_report
import openpyxl
import csv
import io

app = FastAPI(title="BenchmarkIQ API", version="1.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://localhost:8080"],
    allow_methods=["*"],
    allow_headers=["*"],
)

engine = get_engine()

@app.get("/api/stats")
def get_stats():
    with engine.connect() as conn:
        companies = conn.execute(text("SELECT COUNT(*) FROM companies")).fetchone()[0]
        sectors   = conn.execute(text("SELECT COUNT(DISTINCT sector_name) FROM sectors")).fetchone()[0]
        data_points = conn.execute(text("SELECT COUNT(*) FROM profit_loss")).fetchone()[0]
        avg_score = conn.execute(text("SELECT ROUND(AVG(total_score),1) FROM benchmark_scores")).fetchone()[0]
    return {
        "companies":   companies,
        "sectors":     sectors,
        "data_points": data_points,
        "avg_score":   float(avg_score) if avg_score else 0
    }

@app.get("/api/companies")
def get_companies():
    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT c.company_id, c.company_name, c.bse_code,
                   c.nse_symbol, c.market_cap, s.sector_name,
                   b.total_score, b.industry_rank,
                   r.roe, r.roce, r.revenue_growth, r.net_margin
            FROM companies c
            JOIN sectors s ON c.sector_id = s.sector_id
            LEFT JOIN benchmark_scores b ON b.company_id = c.company_id
                AND b.quarter = 'Q4FY25'
            LEFT JOIN financial_ratios r ON r.company_id = c.company_id
                AND r.quarter = 'Q4FY25'
            ORDER BY c.company_name
        """)).fetchall()

    def f(v): return float(v) if v is not None else 0

    return [
        {
            "company_id":     r[0],
            "company_name":   r[1],
            "bse_code":       r[2],
            "nse_symbol":     r[3],
            "market_cap":     f(r[4]),
            "sector_name":    r[5],
            "total_score":    f(r[6]),
            "rank":           r[7],
            "roe":            f(r[8]),
            "roce":           f(r[9]),
            "revenue_growth": f(r[10]),
            "net_margin":     f(r[11]),
        }
        for r in rows
    ]

@app.get("/api/company/{company_id}")
def get_company(company_id: int):
    with engine.connect() as conn:
        info = conn.execute(text("""
            SELECT c.company_name, c.bse_code, c.nse_symbol,
                   c.market_cap, s.sector_name
            FROM companies c
            JOIN sectors s ON c.sector_id = s.sector_id
            WHERE c.company_id = :id
        """), {"id": company_id}).fetchone()

        if not info:
            raise HTTPException(status_code=404, detail="Company not found")

        ratios = conn.execute(text("""
            SELECT roe, roce, net_margin, debt_to_equity,
                   revenue_growth, cagr_3y, cagr_5y, ebitda_margin
            FROM financial_ratios
            WHERE company_id = :id AND quarter = 'Q4FY25'
        """), {"id": company_id}).fetchone()

        score = conn.execute(text("""
            SELECT total_score, industry_rank, industry_percentile
            FROM benchmark_scores
            WHERE company_id = :id AND quarter = 'Q4FY25'
        """), {"id": company_id}).fetchone()

        trend = conn.execute(text("""
            SELECT quarter, revenue, net_profit, ebitda, eps
            FROM quarterly_results
            WHERE company_id = :id
            ORDER BY period_end DESC LIMIT 10
        """), {"id": company_id}).fetchall()

        pl = conn.execute(text("""
            SELECT fiscal_year, sales, net_profit, interest,
                   depreciation, raw_material, employee_cost, tax
            FROM profit_loss
            WHERE company_id = :id
            ORDER BY fiscal_year
        """), {"id": company_id}).fetchall()

        bs = conn.execute(text("""
            SELECT fiscal_year, equity_capital, reserves,
                   borrowings, total_assets, cash_and_bank,
                   receivables, inventory
            FROM balance_sheet
            WHERE company_id = :id
            ORDER BY fiscal_year
        """), {"id": company_id}).fetchall()

        cf = conn.execute(text("""
            SELECT fiscal_year, operating_cf,
                   investing_cf, financing_cf, net_cash_flow
            FROM cash_flow
            WHERE company_id = :id
            ORDER BY fiscal_year
        """), {"id": company_id}).fetchall()

    def f(v): return float(v) if v is not None else None

    return {
        "info": {
            "name":       info[0],
            "bse_code":   info[1],
            "nse_symbol": info[2],
            "market_cap": f(info[3]),
            "sector":     info[4]
        },
        "ratios": {
            "roe":            f(ratios[0]) if ratios else None,
            "roce":           f(ratios[1]) if ratios else None,
            "net_margin":     f(ratios[2]) if ratios else None,
            "debt_to_equity": f(ratios[3]) if ratios else None,
            "revenue_growth": f(ratios[4]) if ratios else None,
            "cagr_3y":        f(ratios[5]) if ratios else None,
            "cagr_5y":        f(ratios[6]) if ratios else None,
            "ebitda_margin":  f(ratios[7]) if ratios else None,
        } if ratios else {},
        "benchmark": {
            "score":      f(score[0]) if score else None,
            "rank":       score[1]    if score else None,
            "percentile": f(score[2]) if score else None,
        } if score else {},
        "trend": [
            {"quarter": r[0], "revenue": f(r[1]),
             "net_profit": f(r[2]), "ebitda": f(r[3]), "eps": f(r[4])}
            for r in reversed(trend)
        ],
        "pl": [
            {"year": r[0], "sales": f(r[1]), "net_profit": f(r[2]),
             "interest": f(r[3]), "depreciation": f(r[4]),
             "raw_material": f(r[5]), "employee_cost": f(r[6]), "tax": f(r[7])}
            for r in pl
        ],
        "balance_sheet": [
            {"year": r[0], "equity": f(r[1]), "reserves": f(r[2]),
             "borrowings": f(r[3]), "total_assets": f(r[4]),
             "cash": f(r[5]), "receivables": f(r[6]), "inventory": f(r[7])}
            for r in bs
        ],
        "cash_flow": [
            {"year": r[0], "operating": f(r[1]), "investing": f(r[2]),
             "financing": f(r[3]), "net": f(r[4])}
            for r in cf
        ]
    }

@app.get("/api/benchmark/{sector_name}")
def get_benchmark(sector_name: str):
    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT c.company_id, c.company_name, c.market_cap,
                   r.roce, r.revenue_growth, r.net_margin,
                   r.roe, r.debt_to_equity,
                   b.total_score, b.industry_rank
            FROM financial_ratios r
            JOIN companies c ON r.company_id = c.company_id
            JOIN sectors s ON c.sector_id = s.sector_id
            LEFT JOIN benchmark_scores b ON b.company_id = c.company_id
                AND b.quarter = r.quarter
            WHERE s.sector_name = :sector
              AND r.quarter = 'Q4FY25'
        """), {"sector": sector_name}).fetchall()

    def f(v): return float(v) if v is not None else 0

    return [
        {
            "id":             r[0],
            "name":           r[1],
            "market_cap":     f(r[2]),
            "roce":           f(r[3]),
            "revenue_growth": f(r[4]),
            "net_margin":     f(r[5]),
            "roe":            f(r[6]),
            "debt_to_equity": f(r[7]),
            "score":          f(r[8]),
            "rank":           r[9]
        }
        for r in rows
    ]

@app.get("/api/sectors")
def get_sectors():
    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT s.sector_name,
                   COUNT(DISTINCT c.company_id) as company_count,
                   ROUND(AVG(r.roe),1)            as avg_roe,
                   ROUND(AVG(r.roce),1)           as avg_roce,
                   ROUND(AVG(r.net_margin),1)     as avg_margin,
                   ROUND(AVG(r.revenue_growth),1) as avg_growth,
                   ROUND(AVG(b.total_score),1)    as avg_score
            FROM sectors s
            JOIN companies c ON c.sector_id = s.sector_id
            LEFT JOIN financial_ratios r ON r.company_id = c.company_id
                AND r.quarter = 'Q4FY25'
            LEFT JOIN benchmark_scores b ON b.company_id = c.company_id
                AND b.quarter = 'Q4FY25'
            GROUP BY s.sector_name
            ORDER BY avg_score DESC
        """)).fetchall()

    def f(v): return float(v) if v is not None else None

    return [
        {
            "sector_name":   r[0],
            "company_count": r[1],
            "avg_roe":       f(r[2]),
            "avg_roce":      f(r[3]),
            "avg_margin":    f(r[4]),
            "avg_growth":    f(r[5]),
            "avg_score":     f(r[6])
        }
        for r in rows
    ]

@app.get("/api/leaderboard")
def get_leaderboard():
    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT c.company_id, c.company_name, c.nse_symbol,
                   s.sector_name, r.roe, r.roce,
                   r.net_margin, r.revenue_growth, b.total_score
            FROM financial_ratios r
            JOIN companies c ON r.company_id = c.company_id
            JOIN sectors s ON c.sector_id = s.sector_id
            LEFT JOIN benchmark_scores b ON b.company_id = c.company_id
                AND b.quarter = r.quarter
            WHERE r.quarter = 'Q4FY25'
              AND b.total_score IS NOT NULL
            ORDER BY b.total_score DESC
            LIMIT 10
        """)).fetchall()

    def f(v): return float(v) if v is not None else 0

    return [
        {
            "company_id":     r[0],
            "company_name":   r[1],
            "nse_symbol":     r[2],
            "sector_name":    r[3],
            "roe":            f(r[4]),
            "roce":           f(r[5]),
            "net_margin":     f(r[6]),
            "revenue_growth": f(r[7]),
            "total_score":    f(r[8]),
        }
        for r in rows
    ]

def score_label(score):
    if score is None:
        return "N/A"
    if score >= 85: return "Excellence Leader"
    if score >= 70: return "High Performer"
    if score >= 55: return "Above Average"
    if score >= 40: return "Average"
    if score >= 25: return "Below Average"
    return "Needs Improvement"

def percentile_rank(peer_values: dict, target_id, invert=False):
    """peer_values: {id: value}. Returns 0-100 percentile for target_id, or None."""
    items = [(k, v) for k, v in peer_values.items() if v is not None]
    if target_id not in dict(items):
        return None
    items.sort(key=lambda kv: -kv[1] if invert else kv[1])
    n = len(items)
    if n <= 1:
        return 50.0
    rank = next(i for i, (k, _) in enumerate(items) if k == target_id)
    return round((rank / (n - 1)) * 100, 1)

@app.get("/api/reports/{company_id}/pdf")
def get_report_pdf(company_id: int, report_type: str = "Executive Summary", from_date: str = None, to_date: str = None):
    with engine.connect() as conn:
        info = conn.execute(text("""
            SELECT c.company_name, s.sector_name, c.sector_id
            FROM companies c
            JOIN sectors s ON c.sector_id = s.sector_id
            WHERE c.company_id = :id
        """), {"id": company_id}).fetchone()

        if not info:
            raise HTTPException(status_code=404, detail="Company not found")

        ratios_row = conn.execute(text("""
            SELECT roe, roce, net_margin, debt_to_equity, ebitda_margin, revenue_growth
            FROM financial_ratios
            WHERE company_id = :id AND quarter = 'Q4FY25'
        """), {"id": company_id}).fetchone()

        score_row = conn.execute(text("""
            SELECT total_score
            FROM benchmark_scores
            WHERE company_id = :id AND quarter = 'Q4FY25'
        """), {"id": company_id}).fetchone()

        latest_sales = conn.execute(text("""
            SELECT sales FROM profit_loss
            WHERE company_id = :id
            ORDER BY fiscal_year DESC LIMIT 1
        """), {"id": company_id}).fetchone()

        peers = conn.execute(text("""
            SELECT r.company_id, c.company_name, r.roe, r.roce, r.net_margin, r.revenue_growth, r.debt_to_equity
            FROM financial_ratios r
            JOIN companies c ON r.company_id = c.company_id
            WHERE c.sector_id = :sector_id AND r.quarter = 'Q4FY25'
        """), {"sector_id": info[2]}).fetchall()

        yearly = conn.execute(text("""
            SELECT fiscal_year, sales, net_profit
            FROM profit_loss
            WHERE company_id = :id
            ORDER BY fiscal_year
        """), {"id": company_id}).fetchall()

    def f(v):
        return round(float(v), 2) if v is not None else "N/A"

    ratios = {
        "roe":            f(ratios_row[0]) if ratios_row else "N/A",
        "roce":           f(ratios_row[1]) if ratios_row else "N/A",
        "net_margin":     f(ratios_row[2]) if ratios_row else "N/A",
        "debt_to_equity": f(ratios_row[3]) if ratios_row else "N/A",
        "ebitda_margin":  f(ratios_row[4]) if ratios_row else "N/A",
    }

    score = float(score_row[0]) if score_row and score_row[0] is not None else None
    label = score_label(score)
    revenue = f"₹{float(latest_sales[0]):,.2f} Cr" if latest_sales and latest_sales[0] is not None else None

    # column index within each `peers` row tuple, for percentile_rank
    metric_cols = {"roe": 2, "roce": 3, "net_margin": 4, "revenue_growth": 5, "debt_to_equity": 6}
    percentiles, strengths, weaknesses = [], [], []
    for label_name, invert, key in [
        ("ROE", False, "roe"), ("ROCE", False, "roce"), ("Net Margin", False, "net_margin"),
        ("Revenue Growth", False, "revenue_growth"), ("Debt / Equity", True, "debt_to_equity"),
    ]:
        col = metric_cols[key]
        peer_values = {p[0]: (float(p[col]) if p[col] is not None else None) for p in peers}
        pct = percentile_rank(peer_values, company_id, invert=invert)
        if pct is not None:
            percentiles.append({"label": label_name, "pct": pct})
            if pct >= 75:
                strengths.append({"label": label_name, "pct": pct})
            elif pct < 50:
                weaknesses.append({"label": label_name, "pct": pct})
    strengths.sort(key=lambda x: -x["pct"])
    weaknesses.sort(key=lambda x: x["pct"])
    strengths, weaknesses = strengths[:3], weaknesses[:3]

    peer_rows = [{
        "name": p[1], "roe": p[2], "roce": p[3], "net_margin": p[4],
        "revenue_growth": p[5], "debt_to_equity": p[6],
        "is_target": p[0] == company_id,
    } for p in peers]

    yearly_financials = [{
        "year": r[0],
        "sales": float(r[1]) if r[1] is not None else None,
        "net_profit": float(r[2]) if r[2] is not None else None,
    } for r in yearly]

    buffer = generate_pdf_report(
        company_name=info[0],
        sector=info[1],
        ratios=ratios,
        score=score if score is not None else "N/A",
        label=label,
        report_type=report_type,
        revenue=revenue,
        percentiles=percentiles,
        from_date=from_date,
        to_date=to_date,
        yearly_financials=yearly_financials,
        peer_rows=peer_rows,
        strengths=strengths,
        weaknesses=weaknesses,
    )

    filename = f"{info[0].replace(' ', '_')}_{report_type.replace(' ', '_')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

TEMPLATE_COLUMNS = ["Revenue", "EBITDA", "Net Profit", "Total Debt", "Equity", "Capital Employed"]

@app.get("/api/upload/template")
def download_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Company Financials"
    ws.append(TEMPLATE_COLUMNS)
    ws.append([5000, 900, 450, 200, 2000, 2800])  # sample row

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="BenchmarkIQ_Upload_Template.xlsx"'},
    )


def _parse_uploaded_financials(filename: str, raw: bytes) -> dict:
    """Reads the first data row and returns a dict of the 6 template columns. Case-insensitive header match."""
    def normalize(h):
        return str(h).strip().lower().replace(" ", "_")

    wanted = {normalize(c): c for c in TEMPLATE_COLUMNS}

    if filename.lower().endswith(".csv"):
        text_data = raw.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text_data))
        rows = list(reader)
    else:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.active
        rows = [[c for c in row] for row in ws.iter_rows(values_only=True)]

    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="File has no data row below the header.")

    header = [normalize(h) for h in rows[0]]
    data_row = rows[1]

    result = {}
    for i, h in enumerate(header):
        if h in wanted and i < len(data_row):
            val = data_row[i]
            try:
                result[wanted[h]] = float(val)
            except (TypeError, ValueError):
                result[wanted[h]] = None

    missing = [c for c in TEMPLATE_COLUMNS if c not in result or result[c] is None]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing or invalid values for: {', '.join(missing)}. Column names must match the template exactly.")

    return result


@app.post("/api/upload/benchmark")
def upload_benchmark(file: UploadFile = File(...), sector: str = Form(...)):
    raw = file.file.read()
    if len(raw) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File exceeds 10 MB limit.")

    data = _parse_uploaded_financials(file.filename, raw)

    revenue = data["Revenue"]
    ebitda = data["EBITDA"]
    net_profit = data["Net Profit"]
    total_debt = data["Total Debt"]
    equity = data["Equity"]
    capital_employed = data["Capital Employed"]

    if revenue == 0 or equity == 0 or capital_employed == 0:
        raise HTTPException(status_code=400, detail="Revenue, Equity, and Capital Employed must be non-zero.")

    net_margin = round((net_profit / revenue) * 100, 2)
    ebitda_margin = round((ebitda / revenue) * 100, 2)
    roe = round((net_profit / equity) * 100, 2)
    roce_proxy = round((ebitda / capital_employed) * 100, 2)  # EBITDA-based proxy — real ROCE needs EBIT
    debt_to_equity = round(total_debt / equity, 2)

    with engine.connect() as conn:
        sector_check = conn.execute(text("SELECT sector_id FROM sectors WHERE sector_name = :s"), {"s": sector}).fetchone()
        if not sector_check:
            raise HTTPException(status_code=404, detail=f"Sector '{sector}' not found.")

        peers = conn.execute(text("""
            SELECT r.company_id, r.roe, r.roce, r.net_margin, r.debt_to_equity, r.ebitda_margin
            FROM financial_ratios r
            JOIN companies c ON r.company_id = c.company_id
            WHERE c.sector_id = :sector_id AND r.quarter = 'Q4FY25'
        """), {"sector_id": sector_check[0]}).fetchall()

    UPLOAD_ID = -1  # sentinel id, not a real company_id in the DB

    def pct(col_idx, val, invert=False):
        peer_values = {p[0]: (float(p[col_idx]) if p[col_idx] is not None else None) for p in peers}
        peer_values[UPLOAD_ID] = val
        return percentile_rank(peer_values, UPLOAD_ID, invert=invert)

    roe_p = pct(1, roe)
    roce_p = pct(2, roce_proxy)
    margin_p = pct(3, net_margin)
    de_p = pct(4, debt_to_equity, invert=True)
    ebitda_p = pct(5, ebitda_margin)

    ratios_calculated = sum(1 for x in [net_margin, ebitda_margin, roe, roce_proxy, debt_to_equity] if x is not None)

    # Weighted percentile blend — Net Margin 10%, ROE 8%, ROCE(proxy) 5%, D/E 8%, EBITDA Margin 10% (sum 41, normalized)
    parts = [(margin_p, 10), (roe_p, 8), (roce_p, 5), (de_p, 8), (ebitda_p, 10)]
    valid = [(p, w) for p, w in parts if p is not None]
    score = round(sum(p * w for p, w in valid) / sum(w for _, w in valid), 1) if valid else None

    return {
        "ratios_calculated": ratios_calculated,
        "ratios": {
            "net_margin": net_margin,
            "ebitda_margin": ebitda_margin,
            "roe": roe,
            "roce_proxy": roce_proxy,
            "debt_to_equity": debt_to_equity,
            "revenue_growth": None,
            "note": "ROCE is an EBITDA-based proxy (real ROCE requires EBIT). Revenue Growth/CAGR not available from a single-period upload.",
        },
        "score": score,
        "score_label": score_label(score),
        "sector": sector,
        "peer_count": len(peers),
    }

@app.get("/api/compare")
def compare_companies(ids: str):
    id_list = [int(x) for x in ids.split(",")]
    with engine.connect() as conn:
        result = []
        for cid in id_list:
            row = conn.execute(text("""
                SELECT c.company_name, s.sector_name,
                       r.roe, r.roce, r.net_margin,
                       r.debt_to_equity, r.revenue_growth,
                       r.cagr_3y, r.ebitda_margin, b.total_score
                FROM financial_ratios r
                JOIN companies c ON r.company_id = c.company_id
                JOIN sectors s ON c.sector_id = s.sector_id
                LEFT JOIN benchmark_scores b ON b.company_id = c.company_id
                    AND b.quarter = r.quarter
                WHERE r.company_id = :id AND r.quarter = 'Q4FY25'
            """), {"id": cid}).fetchone()

            if row:
                def f(v): return float(v) if v is not None else None
                result.append({
                    "name":           row[0],
                    "sector":         row[1],
                    "roe":            f(row[2]),
                    "roce":           f(row[3]),
                    "net_margin":     f(row[4]),
                    "debt_to_equity": f(row[5]),
                    "revenue_growth": f(row[6]),
                    "cagr_3y":        f(row[7]),
                    "ebitda_margin":  f(row[8]),
                    "score":          f(row[9]),
                })
    return result