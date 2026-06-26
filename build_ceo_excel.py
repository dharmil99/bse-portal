import os, sys
sys.path.insert(0, r'C:\Users\Jignesh\Desktop\bse_portal')

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, BubbleChart, Series
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from sqlalchemy import text
from scripts.db_connect import get_engine

engine = get_engine()

OUTPUT = r'C:\Users\Jignesh\Desktop\bse_portal\CEO_Excellence_Presentation.xlsx'

AUTO_COMPANIES = [
    "Tata Motors Limited", "Maruti Suzuki India", "Mahindra and Mahindra",
    "Bajaj Auto Limited", "Hero MotoCorp Limited", "TVS Motor Company",
    "Eicher Motors Limited", "Ashok Leyland Limited", "Bosch Limited",
    "MRF Limited", "Apollo Tyres Limited", "CEAT Limited",
    "Balkrishna Industries", "Samvardhana Motherson", "Minda Industries Limited",
    "Sona BLW Precision", "Endurance Technologies", "Escorts Kubota Limited",
    "Force Motors Limited", "Atul Auto Limited"
]

RATIO_WEIGHTS = {
    # Profitability = 35%
    "Net Profit Margin":       0.10,
    "EBITDA Margin":           0.10,
    "ROE":                     0.08,
    "ROCE":                    0.05,
    "Operating Profit Margin": 0.02,
    # Growth = 25%
    "Revenue Growth YoY":      0.10,
    "3Y Revenue CAGR":         0.08,
    "NP Growth YoY":           0.05,
    "EPS Growth YoY":          0.02,
    # Efficiency = 20%
    "Asset Turnover":          0.07,
    "Debtor Days":             0.05,
    "Inventory Turnover":      0.08,
    # Safety = 20%
    "Debt to Equity":          0.08,
    "Interest Coverage":       0.07,
    "Current Ratio":           0.05,
}

HIGHER_BETTER = {
    "Net Profit Margin": True, "EBITDA Margin": True,
    "ROE": True, "ROCE": True, "Operating Profit Margin": True,
    "Revenue Growth YoY": True, "3Y Revenue CAGR": True,
    "NP Growth YoY": True, "Asset Turnover": True,
    "Debtor Days": False, "Inventory Turnover": True,
    "Debt to Equity": False, "Interest Coverage": True,
    "EPS Growth YoY": True, "Current Ratio": True,
}

CATEGORY_COLORS = {
    "Profitability": "1F4E79",
    "Growth": "375623",
    "Efficiency": "7030A0",
    "Safety": "843C0C",
}

RATIO_CATEGORIES = {
    "Net Profit Margin": "Profitability", "EBITDA Margin": "Profitability",
    "ROE": "Profitability", "ROCE": "Profitability",
    "Operating Profit Margin": "Profitability",
    "Revenue Growth YoY": "Growth", "3Y Revenue CAGR": "Growth",
    "NP Growth YoY": "Growth", "EPS Growth YoY": "Growth",
    "Asset Turnover": "Efficiency", "Debtor Days": "Efficiency",
    "Inventory Turnover": "Efficiency",
    "Debt to Equity": "Safety", "Interest Coverage": "Safety",
    "Current Ratio": "Safety",
}

# ── Palette ──────────────────────────────────────────────────────────────────
NAVY    = "1F3864"
GOLD    = "C9A84C"
LIGHT_BLUE = "BDD7EE"
DARK_GRAY  = "404040"
MID_GRAY   = "808080"
WHITE      = "FFFFFF"
GREEN_DARK = "375623"
GREEN_LIGHT= "E2EFDA"
RED_LIGHT  = "FFDCDC"
AMBER      = "FFF2CC"

def thin_border():
    s = Side(style='thin', color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def header_font(size=11, bold=True, color=WHITE):
    return Font(name="Arial", size=size, bold=bold, color=color)

def cell_font(size=10, bold=False, color=DARK_GRAY):
    return Font(name="Arial", size=size, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center")

# ── Data fetch (same logic as excellence_model.py) ────────────────────────────
def fetch_company_data(company_name):
    with engine.connect() as conn:
        result = conn.execute(text(
            "SELECT company_id FROM companies WHERE company_name = :name"
        ), {"name": company_name}).fetchone()
        if not result:
            return None, None
        cid = result[0]
        pl = pd.read_sql(text("""
            SELECT fiscal_year, sales, net_profit, raw_material,
                   employee_cost, depreciation, interest,
                   profit_before_tax, other_income
            FROM profit_loss WHERE company_id = :cid ORDER BY fiscal_year
        """), conn, params={"cid": cid})
        bs = pd.read_sql(text("""
            SELECT fiscal_year, equity_capital, reserves, borrowings,
                   total_assets, net_block, receivables, inventory,
                   cash_and_bank, other_liabilities
            FROM balance_sheet WHERE company_id = :cid ORDER BY fiscal_year
        """), conn, params={"cid": cid})
    return pl, bs

def safe_div(a, b, mult=1):
    try:
        if b and float(b) != 0:
            return round(float(a) / float(b) * mult, 2)
    except:
        pass
    return None

def calculate_ratios(pl, bs):
    if pl is None or pl.empty:
        return {}
    merged = pd.merge(pl, bs, on="fiscal_year", how="inner")
    if merged.empty:
        return {}
    latest = merged.iloc[-1]
    prev   = merged.iloc[-2] if len(merged) > 1 else latest
    old3   = merged.iloc[-4] if len(merged) > 3 else merged.iloc[0]
    r = {}
    ebitda = (latest["net_profit"] or 0) + (latest["interest"] or 0) + (latest["depreciation"] or 0)
    equity = (latest.get("equity_capital") or 0) + (latest.get("reserves") or 0)
    cap_emp = equity + (latest.get("borrowings") or 0)
    ebit   = (latest["net_profit"] or 0) + (latest["interest"] or 0)
    r["Net Profit Margin"]       = safe_div(latest["net_profit"], latest["sales"], 100)
    r["EBITDA Margin"]           = safe_div(ebitda, latest["sales"], 100)
    r["ROE"]                     = safe_div(latest["net_profit"], equity, 100)
    r["ROCE"]                    = safe_div(ebit, cap_emp, 100)
    r["Operating Profit Margin"] = safe_div(ebitda, latest["sales"], 100)
    r["Revenue Growth YoY"]      = safe_div(latest["sales"] - prev["sales"], prev["sales"], 100)
    try:
        if old3["sales"] and float(old3["sales"]) != 0:
            r["3Y Revenue CAGR"] = round(((float(latest["sales"]) / float(old3["sales"])) ** (1/3) - 1) * 100, 2)
        else:
            r["3Y Revenue CAGR"] = None
    except:
        r["3Y Revenue CAGR"] = None
    try:
        prev_np = float(prev["net_profit"]); curr_np = float(latest["net_profit"])
        r["NP Growth YoY"] = round((curr_np - prev_np) / abs(prev_np) * 100, 2) if prev_np != 0 else None
    except:
        r["NP Growth YoY"] = None
    r["EPS Growth YoY"] = r["NP Growth YoY"]
    total_assets = float(latest.get("total_assets") or 0)
    if total_assets == 0:
        total_assets = sum(float(latest.get(k) or 0) for k in ["net_block","receivables","inventory","cash_and_bank"])
    r["Asset Turnover"]     = safe_div(latest["sales"], total_assets) if total_assets else None
    r["Debtor Days"]        = safe_div(latest.get("receivables") or 0, latest["sales"], 365)
    r["Inventory Turnover"] = safe_div(latest["sales"], latest.get("inventory")) if latest.get("inventory") else None
    r["Debt to Equity"]     = safe_div(latest.get("borrowings"), equity)
    interest = float(latest.get("interest") or 0)
    r["Interest Coverage"]  = safe_div(ebit, interest) if interest > 0 else None
    curr_assets = sum(float(latest.get(k) or 0) for k in ["receivables","inventory","cash_and_bank"])
    curr_liab   = float(latest.get("other_liabilities") or 0)
    r["Current Ratio"]      = safe_div(curr_assets, curr_liab) if curr_liab > 0 else None
    return r

def percentile_rank(value, all_values, higher_better=True):
    valid = [v for v in all_values if v is not None]
    if not valid or value is None:
        return 50
    if higher_better:
        return round(sum(1 for v in valid if v <= value) / len(valid) * 100, 1)
    else:
        return round(sum(1 for v in valid if v >= value) / len(valid) * 100, 1)

def tier_label(score):
    if score >= 85:   return "Excellence Leader"
    elif score >= 70: return "High Performer"
    elif score >= 55: return "Above Average"
    elif score >= 40: return "Average"
    elif score >= 25: return "Below Average"
    else:             return "Needs Improvement"

def tier_color(score):
    if score >= 85:   return "1F4E79"
    elif score >= 70: return "375623"
    elif score >= 55: return "548235"
    elif score >= 40: return "7F6000"
    elif score >= 25: return "C55A11"
    else:             return "9C0006"

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

# ─────────────────────────────────────────────────────────────────────────────
#  SHEET 1 — COVER
# ─────────────────────────────────────────────────────────────────────────────
def build_cover(wb):
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 20
    for r in range(2, 30):
        ws.row_dimensions[r].height = 24

    # Background fill rows
    for row in range(1, 40):
        for col in range(1, 15):
            ws.cell(row=row, column=col).fill = fill(NAVY)

    def cv(row, col, val, font_size=12, bold=False, color=WHITE, align="center"):
        c = ws.cell(row=row, column=col, value=val)
        c.font  = Font(name="Arial", size=font_size, bold=bold, color=color)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        return c

    # Gold accent bar
    for col in range(1, 15):
        ws.cell(row=6, column=col).fill = fill(GOLD)
    for col in range(1, 15):
        ws.cell(row=7, column=col).fill = fill(GOLD)

    cv(9,  3, "BenchmarkIQ", 36, True, WHITE)
    cv(11, 3, "AUTOMOBILE SECTOR", 22, True, GOLD)
    cv(13, 3, "Excellence Model — CEO Presentation", 16, False, "BDD7EE")
    cv(15, 3, "Comprehensive Financial Performance Analysis", 12, False, "BDD7EE")
    cv(17, 3, "FY2024–25  |  20 Companies  |  15 Key Ratios", 11, False, MID_GRAY)
    cv(20, 3, "CONFIDENTIAL", 10, True, GOLD)
    cv(22, 3, "Prepared by BenchmarkIQ Analytics", 10, False, MID_GRAY)

    ws.merge_cells("C9:L9")
    ws.merge_cells("C11:L11")
    ws.merge_cells("C13:L13")
    ws.merge_cells("C15:L15")
    ws.merge_cells("C17:L17")
    ws.merge_cells("C20:L20")
    ws.merge_cells("C22:L22")

    for col in range(1, 15):
        set_col_width(ws, col, 8)
    set_col_width(ws, 3, 40)

# ─────────────────────────────────────────────────────────────────────────────
#  SHEET 2 — RANKINGS DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────
def build_rankings(wb, ranked, total_scores, all_ratios):
    ws = wb.create_sheet("Rankings Dashboard")
    ws.sheet_view.showGridLines = False

    # Title band
    for col in range(1, 14):
        ws.cell(row=1, column=col).fill = fill(NAVY)
        ws.cell(row=2, column=col).fill = fill(NAVY)
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 36

    t = ws.cell(row=2, column=1, value="🏆  BenchmarkIQ Excellence Rankings — Automobile Sector")
    t.font      = Font(name="Arial", size=16, bold=True, color=WHITE)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A2:M2")

    # Sub-header
    ws.row_dimensions[3].height = 6
    ws.row_dimensions[4].height = 20
    for col in range(1, 14):
        ws.cell(row=4, column=col).fill = fill("D9E1F2")

    headers = ["Rank", "Company", "Excellence Score", "Tier", "Profitability", "Growth", "Efficiency", "Safety"]
    col_widths = [6, 32, 16, 22, 16, 14, 14, 12]
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font      = Font(name="Arial", size=10, bold=True, color=NAVY)
        c.alignment = center()
        set_col_width(ws, i, w)

    # Category sub-scores
    ratio_names = list(RATIO_WEIGHTS.keys())
    all_vals_map = {r: [all_ratios[c].get(r) for c in AUTO_COMPANIES] for r in ratio_names}
    pct_scores = {}
    for company in AUTO_COMPANIES:
        pct_scores[company] = {}
        for r in ratio_names:
            val = all_ratios[company].get(r)
            pct_scores[company][r] = percentile_rank(val, all_vals_map[r], HIGHER_BETTER[r])

    def cat_score(company, cat):
        cat_ratios = [r for r, c in RATIO_CATEGORIES.items() if c == cat]
        vals = [pct_scores[company].get(r, 50) for r in cat_ratios]
        return round(sum(vals) / len(vals), 1) if vals else 50

    # Data rows
    ROW_START = 5
    for i, (company, score) in enumerate(ranked):
        row = ROW_START + i
        ws.row_dimensions[row].height = 22
        bg = WHITE if i % 2 == 0 else "F5F8FF"

        short = company.replace(" Limited", "").replace(" India", "")
        vals = [i+1, short, score, tier_label(score),
                cat_score(company, "Profitability"), cat_score(company, "Growth"),
                cat_score(company, "Efficiency"), cat_score(company, "Safety")]

        for j, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=j, value=v)
            c.fill      = fill(bg)
            c.border    = thin_border()
            c.alignment = center() if j != 2 else left()
            c.font      = cell_font(10, j == 1)

        # Rank medal color
        rank_cell = ws.cell(row=row, column=1)
        if i == 0: rank_cell.font = Font(name="Arial", size=11, bold=True, color="C9A84C")
        elif i == 1: rank_cell.font = Font(name="Arial", size=11, bold=True, color=MID_GRAY)
        elif i == 2: rank_cell.font = Font(name="Arial", size=11, bold=True, color="CD7F32")

        # Score bar via color
        score_cell = ws.cell(row=row, column=3)
        score_cell.font = Font(name="Arial", size=11, bold=True, color=tier_color(score))

        # Tier colored pill
        tier_cell = ws.cell(row=row, column=4)
        tier_cell.fill = fill(tier_color(score) if score >= 70 else ("FFF2CC" if score >= 40 else "FFDCDC"))
        tier_cell.font = Font(name="Arial", size=9, bold=True,
                              color=WHITE if score >= 70 else DARK_GRAY)

    # Conditional formatting on score column (C5:C24)
    score_range = f"C{ROW_START}:C{ROW_START + len(ranked) - 1}"
    ws.conditional_formatting.add(score_range,
        ColorScaleRule(start_type="min", start_color="F8696B",
                       mid_type="percentile", mid_value=50, mid_color="FFEB84",
                       end_type="max", end_color="63BE7B"))

    # Category score conditional formatting
    for col_idx, cat in zip([5, 6, 7, 8], ["Profitability", "Growth", "Efficiency", "Safety"]):
        col_letter = get_column_letter(col_idx)
        rng = f"{col_letter}{ROW_START}:{col_letter}{ROW_START + len(ranked) - 1}"
        ws.conditional_formatting.add(rng,
            ColorScaleRule(start_type="min", start_color="F8696B",
                           mid_type="percentile", mid_value=50, mid_color="FFEB84",
                           end_type="max", end_color="63BE7B"))

    # ── Bar Chart — Top 10 scores ──────────────────────────────────────────
    chart_data_row = ROW_START + len(ranked) + 3

    ws.cell(row=chart_data_row - 1, column=1, value="Chart Data (Top 10)")
    ws.cell(row=chart_data_row - 1, column=1).font = Font(name="Arial", size=8, color=MID_GRAY)

    for i, (company, score) in enumerate(ranked[:10]):
        short = company.replace(" Limited", "").replace(" India", "")
        ws.cell(row=chart_data_row + i, column=1, value=short)
        ws.cell(row=chart_data_row + i, column=2, value=score)

    chart = BarChart()
    chart.type = "bar"
    chart.title = "Top 10 Companies — Excellence Score"
    chart.y_axis.title = "Company"
    chart.x_axis.title = "Score"
    chart.style = 10
    chart.width = 18
    chart.height = 12

    data_ref  = Reference(ws, min_col=2, min_row=chart_data_row, max_row=chart_data_row + 9)
    cats_ref  = Reference(ws, min_col=1, min_row=chart_data_row, max_row=chart_data_row + 9)
    chart.add_data(data_ref)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = "1F3864"

    ws.add_chart(chart, f"J{ROW_START}")


# ─────────────────────────────────────────────────────────────────────────────
#  SHEET 3 — RATIO HEATMAP
# ─────────────────────────────────────────────────────────────────────────────
def build_heatmap(wb, ranked, all_ratios, pct_scores_map):
    ws = wb.create_sheet("Ratio Heatmap")
    ws.sheet_view.showGridLines = False

    ratio_names = list(RATIO_WEIGHTS.keys())

    # Title
    for col in range(1, len(ratio_names) + 4):
        ws.cell(row=1, column=col).fill = fill(NAVY)
        ws.cell(row=2, column=col).fill = fill(NAVY)
    ws.row_dimensions[1].height = 6
    ws.row_dimensions[2].height = 34
    t = ws.cell(row=2, column=1, value="Ratio Heatmap — All Companies vs All Metrics")
    t.font = Font(name="Arial", size=14, bold=True, color=WHITE)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(f"A2:{get_column_letter(len(ratio_names)+3)}2")

    # Category color bar (row 3)
    ws.row_dimensions[3].height = 8
    ws.row_dimensions[4].height = 30
    ws.row_dimensions[5].height = 16

    # Column headers
    ws.cell(row=4, column=1, value="Rank").font = header_font(9, True, NAVY)
    ws.cell(row=4, column=2, value="Company").font = header_font(9, True, NAVY)
    ws.cell(row=4, column=3, value="Score").font = header_font(9, True, NAVY)
    for c in [1,2,3]:
        ws.cell(row=4, column=c).fill = fill("D9E1F2")
        ws.cell(row=4, column=c).alignment = center()

    for j, r in enumerate(ratio_names, 4):
        cat = RATIO_CATEGORIES[r]
        cat_col = CATEGORY_COLORS[cat]
        h = ws.cell(row=4, column=j, value=r.replace(" ", "\n"))
        h.font      = Font(name="Arial", size=8, bold=True, color=WHITE)
        h.fill      = fill(cat_col)
        h.alignment = center()
        # Weight row
        w = ws.cell(row=5, column=j, value=f"{int(RATIO_WEIGHTS[r]*100)}%")
        w.font      = Font(name="Arial", size=7, color=DARK_GRAY)
        w.alignment = center()
        w.fill      = fill("F2F2F2")

    # Data
    ROW_START = 6
    for i, (company, score) in enumerate(ranked):
        row = ROW_START + i
        ws.row_dimensions[row].height = 18
        bg = WHITE if i % 2 == 0 else "F8F8F8"

        short = company.replace(" Limited", "").replace(" India", "")
        ws.cell(row=row, column=1, value=i+1).alignment = center()
        ws.cell(row=row, column=1).fill = fill(bg)
        ws.cell(row=row, column=1).font = cell_font(9, True)

        ws.cell(row=row, column=2, value=short)
        ws.cell(row=row, column=2).fill = fill(bg)
        ws.cell(row=row, column=2).font = cell_font(9)
        ws.cell(row=row, column=2).alignment = left()

        sc = ws.cell(row=row, column=3, value=score)
        sc.font = Font(name="Arial", size=9, bold=True, color=tier_color(score))
        sc.fill = fill(bg)
        sc.alignment = center()

        for j, rn in enumerate(ratio_names, 4):
            val = all_ratios[company].get(rn)
            pct = pct_scores_map[company].get(rn, 50)
            display = round(val, 1) if val is not None else "-"
            c = ws.cell(row=row, column=j, value=display)
            c.font      = Font(name="Arial", size=8)
            c.alignment = center()
            c.border    = thin_border()
            # Heat color: green=high pct, red=low
            if val is not None:
                if pct >= 75:   bg_h = "C6EFCE"
                elif pct >= 50: bg_h = "FFEB9C"
                else:           bg_h = "FFC7CE"
                c.fill = fill(bg_h)
            else:
                c.fill = fill("F2F2F2")

    # Column widths
    set_col_width(ws, 1, 6)
    set_col_width(ws, 2, 28)
    set_col_width(ws, 3, 8)
    for j in range(4, len(ratio_names) + 4):
        set_col_width(ws, j, 10)


# ─────────────────────────────────────────────────────────────────────────────
#  SHEET 4 — CATEGORY DEEP DIVE
# ─────────────────────────────────────────────────────────────────────────────
def build_category_deep_dive(wb, ranked, all_ratios, pct_scores_map):
    ws = wb.create_sheet("Category Deep Dive")
    ws.sheet_view.showGridLines = False

    for col in range(1, 30):
        ws.cell(row=1, column=col).fill = fill(NAVY)
        ws.cell(row=2, column=col).fill = fill(NAVY)
    ws.row_dimensions[2].height = 34
    t = ws.cell(row=2, column=1, value="Category Deep Dive — Profitability · Growth · Efficiency · Safety")
    t.font = Font(name="Arial", size=14, bold=True, color=WHITE)
    t.alignment = Alignment(horizontal="left", vertical="center")

    categories = {
        "Profitability": ["Net Profit Margin", "EBITDA Margin", "ROE", "ROCE", "Operating Profit Margin"],
        "Growth":        ["Revenue Growth YoY", "3Y Revenue CAGR", "NP Growth YoY", "EPS Growth YoY"],
        "Efficiency":    ["Asset Turnover", "Debtor Days", "Inventory Turnover"],
        "Safety":        ["Debt to Equity", "Interest Coverage", "Current Ratio"],
    }

    start_col = 1
    ROW_HDR = 4
    ROW_DATA = 6

    for cat, ratios in categories.items():
        cat_col = CATEGORY_COLORS[cat]
        n = len(ratios)

        # Category header spanning ratios + company col
        total_cols = n + 1
        end_col = start_col + total_cols - 1

        # Merge category label
        ws.cell(row=ROW_HDR, column=start_col, value=cat)
        ws.cell(row=ROW_HDR, column=start_col).fill = fill(cat_col)
        ws.cell(row=ROW_HDR, column=start_col).font = Font(name="Arial", size=11, bold=True, color=WHITE)
        ws.cell(row=ROW_HDR, column=start_col).alignment = center()
        if total_cols > 1:
            ws.merge_cells(start_row=ROW_HDR, start_column=start_col,
                           end_row=ROW_HDR, end_column=end_col)
        ws.row_dimensions[ROW_HDR].height = 22

        # Company col header
        ws.cell(row=ROW_HDR + 1, column=start_col, value="Company")
        ws.cell(row=ROW_HDR + 1, column=start_col).fill = fill("D9E1F2")
        ws.cell(row=ROW_HDR + 1, column=start_col).font = cell_font(9, True)
        ws.cell(row=ROW_HDR + 1, column=start_col).alignment = center()
        set_col_width(ws, start_col, 24)

        for k, rn in enumerate(ratios):
            col = start_col + 1 + k
            h = ws.cell(row=ROW_HDR + 1, column=col, value=rn.replace(" ", "\n"))
            h.fill = fill(cat_col)
            h.font = Font(name="Arial", size=8, bold=True, color=WHITE)
            h.alignment = center()
            set_col_width(ws, col, 10)
        ws.row_dimensions[ROW_HDR + 1].height = 28

        # Data
        for i, (company, score) in enumerate(ranked):
            row = ROW_DATA + i
            ws.row_dimensions[row].height = 18
            bg = WHITE if i % 2 == 0 else "F8F8F8"
            short = company.replace(" Limited", "").replace(" India", "")
            c = ws.cell(row=row, column=start_col, value=short)
            c.fill = fill(bg); c.font = cell_font(9); c.alignment = left()
            c.border = thin_border()

            for k, rn in enumerate(ratios):
                col = start_col + 1 + k
                val = all_ratios[company].get(rn)
                pct = pct_scores_map[company].get(rn, 50)
                display = round(val, 1) if val is not None else "-"
                cell = ws.cell(row=row, column=col, value=display)
                cell.font = Font(name="Arial", size=8)
                cell.alignment = center()
                cell.border = thin_border()
                if val is not None:
                    if pct >= 75:   cell.fill = fill("C6EFCE")
                    elif pct >= 50: cell.fill = fill("FFEB9C")
                    else:           cell.fill = fill("FFC7CE")
                else:
                    cell.fill = fill("F2F2F2")

        start_col = end_col + 2  # gap between categories


# ─────────────────────────────────────────────────────────────────────────────
#  SHEET 5 — TOP 5 COMPANY PROFILES
# ─────────────────────────────────────────────────────────────────────────────
def build_top5_profiles(wb, ranked, all_ratios, pct_scores_map):
    ws = wb.create_sheet("Top 5 Profiles")
    ws.sheet_view.showGridLines = False

    ratio_names = list(RATIO_WEIGHTS.keys())

    for col in range(1, 20):
        ws.cell(row=1, column=col).fill = fill(NAVY)
        ws.cell(row=2, column=col).fill = fill(NAVY)
    ws.row_dimensions[2].height = 34
    t = ws.cell(row=2, column=1, value="Top 5 Company Profiles — Detailed Ratio Breakdown")
    t.font = Font(name="Arial", size=14, bold=True, color=WHITE)
    t.alignment = Alignment(horizontal="left", vertical="center")

    ROW_START = 4
    COL_PER_COMPANY = 3
    GAP = 1

    for idx, (company, score) in enumerate(ranked[:5]):
        col_start = 1 + idx * (COL_PER_COMPANY + GAP)

        # Company header
        hdr_row = ROW_START
        ws.row_dimensions[hdr_row].height = 28
        c = ws.cell(row=hdr_row, column=col_start,
                    value=f"#{idx+1}  {company.replace(' Limited','').replace(' India','')}")
        c.fill = fill(tier_color(score))
        c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        c.alignment = center()
        ws.merge_cells(start_row=hdr_row, start_column=col_start,
                       end_row=hdr_row, end_column=col_start + COL_PER_COMPANY - 1)

        # Score row
        ws.row_dimensions[hdr_row + 1].height = 20
        sc = ws.cell(row=hdr_row + 1, column=col_start,
                     value=f"Excellence Score: {score}")
        sc.fill = fill(LIGHT_BLUE)
        sc.font = Font(name="Arial", size=10, bold=True, color=NAVY)
        sc.alignment = center()
        ws.merge_cells(start_row=hdr_row + 1, start_column=col_start,
                       end_row=hdr_row + 1, end_column=col_start + COL_PER_COMPANY - 1)

        # Column sub-headers
        ws.row_dimensions[hdr_row + 2].height = 16
        for j, lbl in enumerate(["Metric", "Value", "Pct"]):
            c = ws.cell(row=hdr_row + 2, column=col_start + j, value=lbl)
            c.fill = fill("D9E1F2")
            c.font = cell_font(8, True, NAVY)
            c.alignment = center()

        # Ratios
        for k, rn in enumerate(ratio_names):
            row = hdr_row + 3 + k
            ws.row_dimensions[row].height = 16
            val = all_ratios[company].get(rn)
            pct = pct_scores_map[company].get(rn, 50)
            bg = WHITE if k % 2 == 0 else "F5F8FF"

            lbl_cell = ws.cell(row=row, column=col_start, value=rn)
            lbl_cell.font = cell_font(8); lbl_cell.fill = fill(bg)
            lbl_cell.alignment = left(); lbl_cell.border = thin_border()

            val_cell = ws.cell(row=row, column=col_start + 1,
                               value=round(val, 1) if val is not None else "-")
            val_cell.font = cell_font(8, True)
            val_cell.alignment = center(); val_cell.border = thin_border()
            if val is not None:
                if pct >= 75:   val_cell.fill = fill("C6EFCE")
                elif pct >= 50: val_cell.fill = fill("FFEB9C")
                else:           val_cell.fill = fill("FFC7CE")
            else:
                val_cell.fill = fill("F2F2F2")

            pct_cell = ws.cell(row=row, column=col_start + 2, value=f"{pct}th")
            pct_cell.font = cell_font(8)
            pct_cell.fill = fill(bg); pct_cell.alignment = center()
            pct_cell.border = thin_border()

        # Column widths
        set_col_width(ws, col_start, 22)
        set_col_width(ws, col_start + 1, 9)
        set_col_width(ws, col_start + 2, 7)


# ─────────────────────────────────────────────────────────────────────────────
#  SHEET 6 — METHODOLOGY
# ─────────────────────────────────────────────────────────────────────────────
def build_methodology(wb):
    ws = wb.create_sheet("Methodology")
    ws.sheet_view.showGridLines = False

    for col in range(1, 10):
        ws.cell(row=1, column=col).fill = fill(NAVY)
        ws.cell(row=2, column=col).fill = fill(NAVY)
    ws.row_dimensions[2].height = 34
    t = ws.cell(row=2, column=1, value="Methodology — How Excellence Scores Are Calculated")
    t.font = Font(name="Arial", size=14, bold=True, color=WHITE)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A2:I2")

    rows = [
        (4,  "OVERVIEW", NAVY, WHITE, True),
        (5,  "The Excellence Model ranks 20 Indian automobile companies using 15 financial ratios across 4 categories.", DARK_GRAY, None, False),
        (6,  "Each company receives a percentile rank (0-100) per ratio, then a weighted average score is computed.", DARK_GRAY, None, False),
        (8,  "SCORING METHODOLOGY", NAVY, WHITE, True),
        (9,  "Step 1: Calculate 15 ratios from P&L and Balance Sheet data for each company (latest FY).", DARK_GRAY, None, False),
        (10, "Step 2: Rank each company's ratio against all 20 peers using percentile ranking.", DARK_GRAY, None, False),
        (11, "Step 3: Apply category weights and compute a weighted Excellence Score (0-100).", DARK_GRAY, None, False),
        (13, "TIER CLASSIFICATION", NAVY, WHITE, True),
        (14, "85-100  →  Excellence Leader   |  70-84  →  High Performer   |  55-69  →  Above Average", DARK_GRAY, None, False),
        (15, "40-54   →  Average              |  25-39  →  Below Average    |  0-24   →  Needs Improvement", DARK_GRAY, None, False),
        (17, "DATA SOURCE", NAVY, WHITE, True),
        (18, "BSE/NSE filings via Screener.in  |  Stored in PostgreSQL  |  FY ending March 2025", DARK_GRAY, None, False),
        (20, "COLOR CODING (HEATMAP)", NAVY, WHITE, True),
        (21, "Green  =  Top quartile (75th percentile+)", "375623", None, False),
        (22, "Yellow =  Middle (50th–75th percentile)", "7F6000", None, False),
        (23, "Red    =  Bottom half (below 50th percentile)", "9C0006", None, False),
    ]

    set_col_width(ws, 1, 80)
    for row, text, color, bg, bold in rows:
        ws.row_dimensions[row].height = 22
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(name="Arial", size=10, bold=bold, color=color if not bold else WHITE)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if bold:
            c.fill = fill(color)
        elif bg:
            c.fill = fill(bg)

    # Weights table
    ws.row_dimensions[25].height = 22
    c = ws.cell(row=25, column=1, value="RATIO WEIGHTS")
    c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    c.fill = fill(NAVY)

    for i, (rn, wt) in enumerate(RATIO_WEIGHTS.items()):
        row = 26 + i
        ws.row_dimensions[row].height = 18
        cat = RATIO_CATEGORIES[rn]
        cat_c = CATEGORY_COLORS[cat]
        c1 = ws.cell(row=row, column=1, value=f"  {rn}  ({cat})  →  {int(wt*100)}%")
        c1.font = Font(name="Arial", size=9, color=DARK_GRAY)
        c1.fill = fill(WHITE if i % 2 == 0 else "F5F8FF")
        c1.border = thin_border()


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    print("📊 Building CEO Presentation Excel...")
    print("=" * 55)

    # Gather data
    all_ratios = {}
    for company in AUTO_COMPANIES:
        print(f"  ↳ {company}")
        pl, bs = fetch_company_data(company)
        all_ratios[company] = calculate_ratios(pl, bs)

    ratio_names = list(RATIO_WEIGHTS.keys())
    all_vals_map = {r: [all_ratios[c].get(r) for c in AUTO_COMPANIES] for r in ratio_names}

    pct_scores_map = {}
    for company in AUTO_COMPANIES:
        pct_scores_map[company] = {}
        for r in ratio_names:
            val = all_ratios[company].get(r)
            pct_scores_map[company][r] = percentile_rank(val, all_vals_map[r], HIGHER_BETTER[r])

    total_scores = {}
    for company in AUTO_COMPANIES:
        total_scores[company] = round(sum(
            pct_scores_map[company].get(r, 50) * w for r, w in RATIO_WEIGHTS.items()
        ), 1)

    ranked = sorted(total_scores.items(), key=lambda x: x[1], reverse=True)

    # Build workbook
    wb = Workbook()
    wb.remove(wb.active)

    print("\n📄 Building sheets...")
    build_cover(wb)
    print("  ✅ Cover")
    build_rankings(wb, ranked, total_scores, all_ratios)
    print("  ✅ Rankings Dashboard")
    build_heatmap(wb, ranked, all_ratios, pct_scores_map)
    print("  ✅ Ratio Heatmap")
    build_category_deep_dive(wb, ranked, all_ratios, pct_scores_map)
    print("  ✅ Category Deep Dive")
    build_top5_profiles(wb, ranked, all_ratios, pct_scores_map)
    print("  ✅ Top 5 Profiles")
    build_methodology(wb)
    print("  ✅ Methodology")

    wb.save(OUTPUT)
    print(f"\n✅ Saved → {OUTPUT}")

    # Print rankings summary
    print("\n🏆 FINAL RANKINGS")
    print("=" * 55)
    for i, (company, score) in enumerate(ranked):
        print(f"  {i+1:2}. {company:<32} {score:5.1f}  {tier_label(score)}")

if __name__ == "__main__":
    main()